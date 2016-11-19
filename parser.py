# -*- coding: utf-8 -*-
import requests
from configparser import ConfigParser
from lxml import html
from openpyxl import load_workbook
import os
from datetime import datetime
from shutil import copyfile
from multiprocessing.dummy import Pool as ThreadPool


TITLE = ['ID: ', ' Заголовок:', ' Цена:', ' Город размещения:', ' Дата размещения', ' Ссылка на товар: ',
         ' Ссылка на изображение:', 'Описание:']
FPATH = 'storage.xlsx'
conf_file = 'parser.ini'
URL = 'https://www.avito.ru/moskva/noutbuki?q=t430'
PROXY = {'http': 'http://10.57.254.103:8080', 'ftp': 'ftp://10.57.254.103:8080', 'https': 'http://10.57.254.103:8080'}
PAGES = 2
proxy = PROXY


def get_config(conf_file):
    config = ConfigParser()
    config.read(conf_file)
    # http_proxy = config.get('general', 'http_proxy')
    # https_proxy = config.get('general', 'https_proxy')
    url = config.get('general', 'url')
    pages = config.get('general', 'pages')
    backup = config.get('general', 'backup')
    new_file = config.get('general', 'new')
    description = config.get('general', 'description')
    return url, int(pages), int(backup), int(new_file), description


def backup_existing_file(path):
    if os.path.exists(path):
        try:
            copyfile(path,
                     path[:-4] + str(datetime.today())[:16].replace(':', '').replace('-', "").replace(" ", '') + path[
                                                                                                                 -5:])
            print('Делаю бекап существующего файла', path)
        except:
            backup_existing_file(path)


def read_xls(full_filename):
    table = []
    wb = load_workbook(full_filename)
    ws = wb.active
    for row in ws.iter_rows():
        table_row = []
        for col in row:
            table_row.append(col.value)
        table.append(table_row)
    print("Файл считан")
    return table


def list_to_dict(project_list):
    project_dict = {}
    for row in project_list:
        project_dict[row[0]] = row[1:]
    return project_dict


def get_html(try_url, proxy, retry=5):
    while retry:
        try:
            response = requests.get(try_url, proxies=proxy)
            print('Получил страницу -', try_url)
            return html.document_fromstring(response.content)
        except:
            print('Не удалось получить страницу по ссылке', try_url, 'Пробую еще раз...')
            retry -= 1
    print('Попытки исчерпаны')
    pass


def get_next_url(url, count):
    try:
        index_sign = url.index('?')
    except ValueError:
        return url + '?p=' + str(count)
    return (url[:index_sign] + '?p=' + str(count) + '&' + url[(index_sign + 1):])


def get_row_table(url, proxy):
    rows_table = []
    doc = get_html(url, proxy)
    if doc is None: pass  # Если страница не получена то выходим с возвратом None
    items = doc.cssselect('div.js-catalog_after-ads .item_table')  # отсекаем всЁ до нужного нам раздела
    i = len(items)
    for item in items:
        id_item = int(item.get('id')[1:])  # узнаём ID товара
        href = 'https://www.avito.ru' + item.cssselect('div.description h3 a')[0].get('href')  # узнаём ссылку на товар
        title = item.cssselect('div.description h3 a')[0].get('title')  # узнаём заголовок объявления
        try:
            src = item.cssselect('div.b-photo a img')[0].get('src')  # узнаём ссылку на картинку для объявления
            if src[:5] != 'http:': src = 'http:' + src
        except:
            src = None

        try:
            price = (item.cssselect('div.description .about')[0].text)  # узнаём цену товара
            price = int(price.replace('\n', '').replace('руб.', '').replace(' ', ''))  # отсекаем лишние символы
        except:
            price = int(str('0'))

        try:
            podrazdel = str(item.cssselect('div.description > div.data > p:nth-child(1)')[0].text)
        except:
            podrazdel = 'Не удалось определить'  # узнаём подраздел или город

        try:
            city = item.cssselect('div.description div.data p:nth-child(2)')[0].text
            # узнаём город в котором продаётся товар
        except:
            city = podrazdel  # город не обнаружен, значит раздела нет,
            # а в переменную для раздела записан город, поэтому присваеваем
            # городу  значение, которое получил раздел,а
            podrazdel = 'Не удалось определить'  # разделу приваеваем значение 'Не удалось определить'

        try:
            # date_item = ''
            date_item = str((item.cssselect('.item_table .data .date')[0].text).replace('\n', '').replace(' ', ''))
            # узнаём дату публикации
            if date_item == None: date_item = ''
        except:
            date_item = ''

        rows_table.append([
            id_item
            , title
            , price
            , city
            , date_item
            , href
            , src
            # , ''
        ])
    return rows_table


def get_table(url, proxy, pages):
    project = []
    print('1.', end='')
    project.extend(get_row_table(url, proxy))
    for i in range(2, pages + 1):
        next_url = get_next_url(url, i)
        page = get_row_table(next_url, proxy)
        print(i, '.', end='')
        if not page: print("Страница пустая. Заканчиваем", '\n')
        project.extend(page)
    return project


def get_description(url, proxy=PROXY):
    descripion_item = ''
    try:
        descripion_items = get_html(url, proxy).cssselect(
            'body > div.item-view-page-layout.item-view-page-layout_content > div.l-content.clearfix > div.item-view > div.item-view-content > div.item-view-left > div.item-view-main.js-item-view-main > div.item-view-block > div > div > p'
        )
        for item in descripion_items: descripion_item = descripion_item + str(item.text_content())
    except:
        descripion_item = ''
    return descripion_item


def get_table_wo_desc(new_project):
    table_wo_desc = []
    i = 0
    for i in range(len(new_project) - 1):
        next = 0
        while not next:
            try:
                if len(new_project[i][7]) == 0:  raise
                next = 1
            except:
                if i >= len(new_project): return table_wo_desc, new_project
                table_wo_desc.append(new_project[i])
                new_project.remove(new_project[i])
    return table_wo_desc, new_project


def add_description(new_project, proxy):
    for i in range(len(new_project)):
        print('Осталось', len(new_project) - 1, end='')
        try:
            if len(new_project[i][7]) != 0:  continue
        except:
            descript = str(get_description(new_project[i][5], proxy))
            new_project[i].append(descript)
        new_project[i][7] = descript
    return new_project


def get_t_desc(table_wo_desc):
    urls = []
    for row in table_wo_desc: urls.append(row[5])

    # Make the Pool of workers
    pool = ThreadPool(8)

    # Open the urls in their own threads
    # and return the results
    results = pool.map(get_description, urls)

    # close the pool and wait for the work to finish
    pool.close()
    pool.join()
    return results


def get_table_with_desc(table_wo_desc, desc_table):
    for i in range(len(table_wo_desc)):
        table_wo_desc[i].append(str(desc_table[i]))
    return table_wo_desc


def xls_write(project, full_filename):
    # with open(full_filename, 'wb') as wb:
    wb = load_workbook(full_filename)
    ws = wb.active
    for row in range(2, len(project)):
        # Первой и третьей ячейке присваеваем формат числовой
        ws.cell(row=row + 1, column=1).data_type = 'n'
        ws.cell(row=row + 1, column=1).value = project[row][0]
        # второй ячейке присваеваем гиперссылку из 6-ой ячейки
        ws.cell(row=row + 1, column=2).hyperlink = project[row][5]
        ws.cell(row=row + 1, column=2).value = project[row][1]
        # Третьей ячейке присваеваем формат числовой
        ws.cell(row=row + 1, column=3).data_type = 'n'
        ws.cell(row=row + 1, column=3).value = project[row][2]
        #
        ws.cell(row=row + 1, column=4).value = project[row][3]
        #
        ws.cell(row=row + 1, column=5).value = project[row][4]
        #
        ws.cell(row=row + 1, column=6).value = project[row][5]
        #
        ws.cell(row=row + 1, column=7).value = project[row][6]
        #
        ws.cell(row=row + 1, column=8).value = project[row][7]
    wb.save(full_filename)
    print('Файл успешно сохранён!')


def dict_to_list(input_dict):
    dictlist = []
    for key in input_dict:
        tmp = []
        tmp.append(key)
        for i in input_dict[key]:  tmp.append(i)
        dictlist.append(tmp)
    return dictlist


def main():
    path = FPATH
    url, pages, backup_file, new_file, description = get_config(
        conf_file)  # print("Прочитали и спарсили конфиг", conf_file)
    if backup_file:     backup_existing_file(path)
    if new_file:
        os.remove(path)
        copyfile('storage_template.xlsx', path)
    old_data_list = read_xls(path)  # print('Прочитали файл', path)
    try:
        old_data_list.remove(TITLE)
    except:
        qqq = 'Делать ничего не нужно, т.к. заголовка уже нет.'
    old_data_dict = list_to_dict(old_data_list)  # print("Преобразовали список в словарь.")
    new_data_list = get_table(url, proxy, pages)  # print("Получили все необходимые страницы")
    new_data_dict = list_to_dict(new_data_list)  # print("Преобразовали полученный список в словарь")
    old_data_dict.update(new_data_dict)  # print("Совместили словари, исключив повторения")
    new_data_list = dict_to_list(old_data_dict)  # print("Преобразовали словарь обратно в список")
    # new_data_list = [[874133637, 'Нетбук EM350-21G16I в Москве', 4000, 'м.\xa0Новокосино', 'Сегодня11:04', 'https://www.avito.ru/moskva/noutbuki/netbuk_em350-21g16i_874133637', 'http://94.img.avito.st/140x105/3171007394.jpg', 'EMACHINES EM350-21G16I ЧЕРНЫЙПроцессор Intel Atom N450 1660 MhzКоличество ядер 2Экран 10.1 дюймаОбъем оперативной памяти 2 ГбЖЕСТКИЙ ДИСК (HDD) 160 ГБ Wi-Fi Кард-ридер Микрофон Web-камераБАТАРЕЯ     3 часа 20 минОТЛИЧНОЕ СОСТОЯНИЕ'], [874266764, 'Ноутбук Futjizu 17.3" i7 8ядер 1000гб мощный немец в Москве', 30000, 'м.\xa0Цветной бульвар', 'Сегодня14:32', 'https://www.avito.ru/moskva/noutbuki/noutbuk_futjizu_17.3_i7_8yader_1000gb_moschnyy_nemets_874266764', 'http://81.img.avito.st/140x105/3171490481.jpg', 'Для всего! ФутжицуНемецкое качество! Ноутбук игровой 17,3"Процессор core i7 8ядер Windows82,4gz. 8gb память Hd graphics 4000 Nvidia GeForce 640m le1000 жёсткий диск Матовый экран Супер звук, turboboost, корпус и экран без дефектов'], [874129828, 'Asus N571J в Москве', 60000, 'м.\xa0Полежаевская', 'Сегодня11:00', 'https://www.avito.ru/moskva/noutbuki/asus_n571j_874129828', 'http://57.img.avito.st/140x105/3170998357.jpg', 'Игровой ноутбук Asus N751J. На Windows 10. Корпус матовый алюминий серого цвета. Размеры 41.6 x 3.7 x 28.3 см. Вес  3 кг.Процессор ноутбука\tIntel Core i7 4710HQ 2.5 ГГц; Turbo Boost 3.5 ГГц. Диагональю 17.3 (43.9 см), подсветка Светодиодная (LED). Тип видео\tДискретноеВидео ноутбука\tGeForce GTX 850M (128 бит)Картридер ноутбука\tSDXC, SDHC, SD, MMC Оперативная память 4 гб, плюс 2гб дополнительно поставил. Покупал два месяца назад за 85 тыс. Коробка, чек, гарантия все есть, отдаю полный комплект. СРОЧНО звонить до 24 часов'], [874243916, 'Аккумулятор для ноутбука HP в Москве', 900, 'м.\xa0Братиславская', 'Сегодня13:55', 'https://www.avito.ru/moskva/noutbuki/akkumulyator_dlya_noutbuka_hp_874243916', 'http://53.img.avito.st/140x105/3171398353.jpg', None], [874273022, 'Lenovo Y510P отличный игровой ноутбук в Москве', 37000, 'м.\xa0Новокосино', 'Сегодня14:41', 'https://www.avito.ru/moskva/noutbuki/lenovo_y510p_otlichnyy_igrovoy_noutbuk_874273022', 'http://90.img.avito.st/140x105/3171502390.jpg', None], [853655061, 'Оперативная память MacBook Pro 2шт по 1GB pc3-8500 в Москве', 500, 'м.\xa0Тимирязевская', 'Сегодня15:17', 'https://www.avito.ru/moskva/tovary_dlya_kompyutera/operativnaya_pamyat_macbook_pro_2sht_po_1gb_pc3-8500_853655061', 'http://92.img.avito.st/140x105/3079363192.jpg', None], [874271767, 'Состоянии на 5 + новый в Москве', 7600, 'м.\xa0Щелковская', 'Сегодня14:39', 'https://www.avito.ru/moskva/noutbuki/sostoyanii_na_5_novyy_874271767', 'http://80.img.avito.st/140x105/3171501180.jpg', None], [874132388, 'Fujitsu Siemens Amilo Pro V3405 (доставка) в Москве', 4000, 'Не удалось определить', 'Сегодня11:03', 'https://www.avito.ru/moskva/noutbuki/fujitsu_siemens_amilo_pro_v3405_dostavka_874132388', 'http://05.img.avito.st/140x105/3171005505.jpg', 'Fujitsu Siemens Amilo Pro V3405. Полный комплект. Доставка по Москве бесплатно.'], [874132442, 'Macbook 13 retina 512 gb в Москве', 13000, 'м.\xa0Деловой центр', 'Сегодня11:03', 'https://www.avito.ru/moskva/noutbuki/macbook_13_retina_512_gb_874132442', 'http://15.img.avito.st/140x105/3171006115.jpg', 'Рст, срочно, все вопросы в отклик авито'], [750445340, 'Работай, играй, отдыхай. Lenovo B серия в Москве', 15000, 'м.\xa0Рязанский проспект', 'Сегодня11:02', 'https://www.avito.ru/moskva/noutbuki/rabotay_igray_otdyhay._lenovo_b_seriya_750445340', 'http://58.img.avito.st/140x105/3051350158.jpg', None], [874267354, 'Игровой Core i5 8гб/1000гб+ 2 видюхи+ гарантия в Москве', 26500, 'м.\xa0Павелецкая', 'Сегодня14:32', 'https://www.avito.ru/moskva/noutbuki/igrovoy_core_i5_8gb1000gb_2_vidyuhi_garantiya_874267354', 'http://77.img.avito.st/140x105/3171483877.jpg', None], [874271007, 'Ноутбук в Москве', 8000, 'м.\xa0Кантемировская', 'Сегодня14:38', 'https://www.avito.ru/moskva/noutbuki/noutbuk_874271007', 'http://15.img.avito.st/140x105/3171471215.jpg', 'Продаю срочно......'], [695809122, '17.3"Lenovo -Core i3 + 4Gb + 500Gb + GeForce GT745 в Москве', 25990, 'м.\xa0Рязанский проспект', 'Сегодня11:00', 'https://www.avito.ru/moskva/noutbuki/17.3lenovo_-core_i3_4gb_500gb_geforce_gt745_695809122', 'http://64.img.avito.st/140x105/2126504664.jpg', None], [874272996, 'Рабочая лошадка HP ProBook в Москве', 9200, 'м.\xa0Павелецкая', 'Сегодня14:41', 'https://www.avito.ru/moskva/noutbuki/rabochaya_loshadka_hp_probook_874272996', 'http://84.img.avito.st/140x105/3171504184.jpg', 'Достойная версия пробуков от HP, работали работают и будут работать, оптовые продажи тоже есть, если вдруг) Характеристики :Процессор - Атлон X2 Rм-76-2,30GHzВидеокарта - Radeon  4330 Оперативка - 3ГигобайтаЖесткий - 320гигабайтДиагональ -15.6Win7//WiFi/BT/Web-CamАккумулятор от 30 минут  автономной работы'], [874272421, 'Асус ноутбук в Москве', 7500, 'м.\xa0Лермонтовский проспект', 'Сегодня14:40', 'https://www.avito.ru/moskva/noutbuki/asus_noutbuk_874272421', 'http://24.img.avito.st/140x105/3171492424.jpg', None], [874272497, 'Король Lenovo ThinkPad T430 Core i5 3320M 256 SSD в Москве', 25000, 'Не удалось определить', 'Сегодня14:41', 'https://www.avito.ru/moskva/noutbuki/korol_lenovo_thinkpad_t430_core_i5_3320m_256_ssd_874272497', 'http://04.img.avito.st/140x105/3171469004.jpg', 'Неубиваемая Машина !!!'], [873714029, 'Оперативная память macbook 2x2gb в Москве', 2000, 'м.\xa0Перово', 'Сегодня14:40', 'https://www.avito.ru/moskva/tovary_dlya_kompyutera/operativnaya_pamyat_macbook_2x2gb_873714029', 'http://72.img.avito.st/140x105/3169534072.jpg', 'Оперативная память, снята при замене на 2 планки по 4gb с макбука 2011 года, оригинальная память самсунг'], [874253085, 'Новый (4ядра) в упаковке 4гб/500гб в Москве', 19000, 'м.\xa0Павелецкая', 'Сегодня14:10', 'https://www.avito.ru/moskva/noutbuki/novyy_4yadra_v_upakovke_4gb500gb_874253085', 'http://76.img.avito.st/140x105/3171434576.jpg', None], [874257200, 'Тонкий, 4 ядерный hp, состояние нового в Москве', 12000, 'Не удалось определить', 'Сегодня14:17', 'https://www.avito.ru/moskva/noutbuki/tonkiy_4_yadernyy_hp_sostoyanie_novogo_874257200', 'http://93.img.avito.st/140x105/3171445793.jpg', None], [836271345, '1gb 2rx8 pc2-5300f-555-11 hynix AB-C 12 шт в Москве', 5000, 'м.\xa0Китай-город', 'Сегодня12:24', 'https://www.avito.ru/moskva/tovary_dlya_kompyutera/1gb_2rx8_pc2-5300f-555-11_hynix_ab-c_12_sht_836271345', 'http://29.img.avito.st/140x105/2977760529.jpg', None], [856585205, 'Мощный на core i3 (металлический) с 4 ядрами в Москве', 15400, 'м.\xa0Павелецкая', 'Сегодня13:01', 'https://www.avito.ru/moskva/noutbuki/moschnyy_na_core_i3_metallicheskiy_s_4_yadrami_856585205', 'http://38.img.avito.st/140x105/3096251038.jpg', '****** ****** ****** ******* ****** ****** ******* *******Даем письменную гарантию на обмен ноутбка,если он не подошел вам после покупки.****** ****** ****** ***** ***** ****** ****** ****** *******'], [873562335, 'Новая озу Kingston KVR16N11S8/4 в Москве', 2500, 'м.\xa0ВДНХ', 'Сегодня11:01', 'https://www.avito.ru/moskva/tovary_dlya_kompyutera/novaya_ozu_kingston_kvr16n11s84_873562335', 'http://51.img.avito.st/140x105/3168986351.jpg', 'Не подошли - высокая частота. В магазине брал за 3000р.'], [874130750, '10.1" красный нетбук Hp mini 3.2ghz/1gb/250gb Hdd в Москве', 6499, 'м.\xa0Сходненская', 'Сегодня11:02', 'https://www.avito.ru/moskva/noutbuki/10.1_krasnyy_netbuk_hp_mini_3.2ghz1gb250gb_hdd_874130750', 'http://56.img.avito.st/140x105/3170997856.jpg', 'Шикарный в чехле , который отдам Красивый красный нетбук10.1" нетбук от HP'], [873716223, 'HP 16Gb 4Rx4 PC3-8500R DDR3-1066 CL7 (500666-B21) в Москве', 13500, 'м.\xa0Петровско-Разумовская', 'Сегодня14:43', 'https://www.avito.ru/moskva/tovary_dlya_kompyutera/hp_16gb_4rx4_pc3-8500r_ddr3-1066_cl7_500666-b21_873716223', 'http://99.img.avito.st/140x105/3169544199.jpg', 'Продаётся оригинальная оперативная память\xa0HP 16Gb 4Rx4 PC3-8500R DDR3-1066 ECC CL7 (500666-B21, 500207-071). Состояние ОЕМ.']]
    t_wo_desc, new_data_list = get_table_wo_desc(new_data_list)

    if description:
        t_descs = get_t_desc(t_wo_desc)
        t_with_desc = get_table_with_desc(t_wo_desc, t_descs)
        new_data_list.extend(t_with_desc)
    new_data_list.insert(0, TITLE)  # print('Добавили заголовок к списку')

    xls_write(new_data_list, path)  # print("Записали изменения в файл", path)

    input("Для выхода нажмите Enter")


if __name__ == '__main__':
    main()


