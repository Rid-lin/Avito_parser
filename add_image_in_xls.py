import os

import requests
from openpyxl import load_workbook
from openpyxl.drawing.image import Image


def add_loc_img(new_project):
    for i in range(1, len(new_project)):
        if not new_project[i][6]:
            loc_filename = 'img\\No_image.png'
        elif str(new_project[i][6]).find('//') == -1:
            new_project[i][6] = None
            loc_filename = 'img\\No_image.png'
        else:
            loc_filename = 'img\\' + new_project[i][6].replace('http://', '_').replace('//', '_').replace('/',
                                                                                                          '_').replace(
                ':', '_')
        # print(loc_filename)
        try:
            new_project[i][7] = loc_filename
        except IndexError:
            new_project[i].append(loc_filename)


def get_loc_img(new_project, proxy):
    for row in new_project[1:]:
        url = row[6]
        filename = row[7]
        if not (url or filename): continue
        if url == 'None': continue
        if os.path.exists(filename): continue
        if url[:5] != 'http:': url = 'http:' + url
        print('url', url, 'filename', filename)
        r = requests.get(url, proxies=proxy)
        with open(filename, 'wb') as fd:
            fd.write(r.content)


def xls_write_with_image(project, full_filename):
    # with open(full_filename, 'wb') as wb:
    wb = load_workbook(full_filename)
    ws = wb.active
    for row in range(2, len(project)):
        cols = len(project[0])
        # Первой и третьей ячейке присваеваем формат числовой
        ws.cell(row=row + 1, column=1).data_type = 'n'
        ws.cell(row=row + 1, column=1).value = project[row][0]
        # второй ячейке присваеваем гиперссылку из 6-ой ячейки
        ws.cell(row=row + 1, column=2).hyperlink = project[row][5]
        ws.cell(row=row + 1, column=2).value = project[row][1]
        # Третьей ячейке присваеваем формат числовой
        ws.cell(row=row + 1, column=3).data_type = 'n'
        ws.cell(row=row + 1, column=3).value = project[row][2]
        #
        ws.cell(row=row + 1, column=4).value = project[row][3]
        #
        ws.cell(row=row + 1, column=5).value = project[row][4]
        #
        ws.cell(row=row + 1, column=6).value = project[row][5]
        #
        ws.cell(row=row + 1, column=7).value = project[row][6]
        #
        ws.cell(row=row + 1, column=8).value = project[row][7]
        try:
            img = Image(project[row][7])
            img.anchor(ws.cell(row=row + 1, column=1))
        except:
            qqq = '1'
        ws.add_image(img)
    wb.save(full_filename)
    print('Файл успешно сохранён!')
