# -*- coding: utf-8 -*-
import requests
from configparser import ConfigParser
from lxml import html
import csv
import os
from openpyxl import load_workbook

FPATH = 'storage.xlsx'
conf_file = 'parser.ini'
URL = 'https://www.avito.ru/moskva/noutbuki?q=t430'
PROXY = {'http': 'http://proxy.loc:8080',
         'https': 'http://proxy.loc:8080'}
PAGES = 2
project = [[
    '106102393;Think Pad 600 IBM в Москве;5000;м.\xa0Теплый стан;Сегодня19:03;https://www.avito.ru/moskva/noutbuki/think_pad_600_ibm_106102393;//72.img.avito.st/140x105/199774072.jpg;Не удалось определить'],
    [
        '869891143;Мощный ультрабук asus в Москве;45000;Не удалось определить;Сегодня19:02;https://www.avito.ru/moskva/noutbuki/moschnyy_ultrabuk_asus_869891143;//28.img.avito.st/140x105/3155476428.jpg;Не удалось определить'],
    [
        '869890401;MacBook Pro i5 256 в Москве;70000;м.\xa0Университет;Сегодня19:01;https://www.avito.ru/moskva/noutbuki/macbook_pro_i5_256_869890401;//32.img.avito.st/140x105/3155471132.jpg;Не удалось определить'],
    [
        '869890300;Надёжный ноутбук ibm в Москве;4500;м.\xa0Савеловская;Сегодня19:00;https://www.avito.ru/moskva/noutbuki/nadezhnyy_noutbuk_ibm_869890300;//77.img.avito.st/140x105/3155472777.jpg;Компания'],
    [
        '869890055;Dell Alienware в Москве;80000;м.\xa0ВДНХ;Сегодня19:00;https://www.avito.ru/moskva/noutbuki/dell_alienware_869890055;//31.img.avito.st/140x105/3155468231.jpg;Не удалось определить'],
    [
        '869889779;Samsung ultrabook 13.3 в идеале в Москве;18500;м.\xa0Ботанический сад;Сегодня18:59;https://www.avito.ru/moskva/noutbuki/samsung_ultrabook_13.3_v_ideale_869889779;//01.img.avito.st/140x105/3155470801.jpg;Не удалось определить'],
    [
        '869888569;Асус а 6000 в Москве;1500;м.\xa0Бульвар Дмитрия Донского;Сегодня18:57;https://www.avito.ru/moskva/noutbuki/asus_a_6000_869888569;//29.img.avito.st/140x105/3155465029.jpg;Не удалось определить'],
    [
        '840365487;Ультрабук Asus X552 без единой царапины на гаранти в Москве;12000;Не удалось определить;Сегодня18:55;https://www.avito.ru/moskva/noutbuki/ultrabuk_asus_x552_bez_edinoy_tsarapiny_na_garanti_840365487;//34.img.avito.st/140x105/3001594834.jpg;Не удалось определить'],
    [
        '832087063;MacBook Air 13 mid 2015 Идеал. i5 1.6Ghz 256 HD600 в Москве;64200;Компания;Сегодня18:54;https://www.avito.ru/moskva/noutbuki/macbook_air_13_mid_2015_ideal._i5_1.6ghz_256_hd600_832087063;//31.img.avito.st/140x105/2953254031.jpg;Не удалось определить'],
    [
        '869884609;Ультрабук Samsung NP530 Core i5 в Москве;19900;м.\xa0Коломенская;Сегодня18:49;https://www.avito.ru/moskva/noutbuki/ultrabuk_samsung_np530_core_i5_869884609;//92.img.avito.st/140x105/3155450992.jpg;Не удалось определить'],
    [
        '797425523;Hp compaq nx6310 не рабочий на запчасти в Москве;1500;м.\xa0Курская;Сегодня18:46;https://www.avito.ru/moskva/noutbuki/hp_compaq_nx6310_ne_rabochiy_na_zapchasti_797425523;//55.img.avito.st/140x105/3155445055.jpg;Компания'],
    [
        '778919170;Ноутбук Sony Vaio в Москве;12000;м.\xa0Отрадное;Сегодня18:46;https://www.avito.ru/moskva/noutbuki/noutbuk_sony_vaio_778919170;//08.img.avito.st/140x105/2643897408.jpg;Компания'],
    [
        '869882879;Нр omnibook в Москве;2500;м.\xa0Выхино;Сегодня18:46;https://www.avito.ru/moskva/noutbuki/nr_omnibook_869882879;//79.img.avito.st/140x105/3155445979.jpg;Компания'],
    [
        '869882426;Ноутбук sony vaio VPC-EB3M1R в Москве;13000;Не удалось определить;Сегодня18:45;https://www.avito.ru/moskva/noutbuki/noutbuk_sony_vaio_vpc-eb3m1r_869882426;//60.img.avito.st/140x105/3155429060.jpg;Не удалось определить'],
    [
        '869882256;Приобрету матрицу в Москве;;Не удалось определить;Сегодня18:45;https://www.avito.ru/moskva/noutbuki/priobretu_matritsu_869882256;Нет картинки;Не удалось определить'],
    [
        '869882104;Нетбук asus в Москве;16000;м.\xa0Пролетарская;Сегодня18:44;https://www.avito.ru/moskva/noutbuki/netbuk_asus_869882104;//59.img.avito.st/140x105/3155441159.jpg;Не удалось определить'],
    [
        '869882028;Плата HP daut1AMB6E1 REV. E 574680-001 для ноутубк в Москве;15000;м.\xa0Нагорная;Сегодня18:44;https://www.avito.ru/moskva/noutbuki/plata_hp_daut1amb6e1_rev._e_574680-001_dlya_noutubk_869882028;//56.img.avito.st/140x105/3155427356.jpg;Не удалось определить'],
    [
        '869881925;MacBook Air 13 в Москве;31500;м.\xa0Сокол;Сегодня18:44;https://www.avito.ru/moskva/noutbuki/macbook_air_13_869881925;//54.img.avito.st/140x105/3155450154.jpg;Не удалось определить'],
    [
        '740504473;Lenovo x201 tablet в Москве;14500;м.\xa0Молодежная;Сегодня18:43;https://www.avito.ru/moskva/noutbuki/lenovo_x201_tablet_740504473;//38.img.avito.st/140x105/2359540238.jpg;Не удалось определить'],
    [
        '869880466;Продам Ноутбук Lenovo ideapad 100-14IBY/100-15IBY в Москве;14000;м.\xa0Щелковская;Сегодня18:42;https://www.avito.ru/moskva/noutbuki/prodam_noutbuk_lenovo_ideapad_100-14iby100-15iby_869880466;//54.img.avito.st/140x105/3155429854.jpg;Не удалось определить'],
    [
        '851740849;Продам ноутбук HP Pavilion dv6000 в Москве;5000;м.\xa0Новые Черемушки;Сегодня18:41;https://www.avito.ru/moskva/noutbuki/prodam_noutbuk_hp_pavilion_dv6000_851740849;//41.img.avito.st/140x105/3084082841.jpg;Не удалось определить'],
    [
        '869879951;Ноутбук в Москве;9000;Компания;Сегодня18:41;https://www.avito.ru/moskva/noutbuki/noutbuk_869879951;//52.img.avito.st/140x105/3155436952.jpg;Не удалось определить'],
    [
        '869878890;Отличник для всего Toshiba Satellite Гарантия/Дост в Москве;9990;м.\xa0Павелецкая;Сегодня18:40;https://www.avito.ru/moskva/noutbuki/otlichnik_dlya_vsego_toshiba_satellite_garantiyadost_869878890;//85.img.avito.st/140x105/3155434885.jpg;None'],
    [
        '869877327;Asus в Москве;9500;Компания;Сегодня18:37;https://www.avito.ru/moskva/noutbuki/asus_869877327;//86.img.avito.st/140x105/3155428486.jpg;Не удалось определить'],
    [
        '869875255;Ноутбук packard bel te69 kb в Москве;13500;м.\xa0Белорусская;Сегодня18:33;https://www.avito.ru/moskva/noutbuki/noutbuk_packard_bel_te69_kb_869875255;//26.img.avito.st/140x105/3155428926.jpg;Не удалось определить'],
    [
        '842492479;Продам Ультра бук asus zenbook UX32VD в Москве;30000;м.\xa0Киевская;Сегодня18:32;https://www.avito.ru/moskva/noutbuki/prodam_ultra_buk_asus_zenbook_ux32vd_842492479;//88.img.avito.st/140x105/3013826288.jpg;Не удалось определить'],
    [
        '869874508;Lenovo g510 (i7/ssd128/ram 8gb/hd 8750 2gb) в Москве;37000;м.\xa0Южная;Сегодня18:31;https://www.avito.ru/moskva/noutbuki/lenovo_g510_i7ssd128ram_8gbhd_8750_2gb_869874508;//70.img.avito.st/140x105/3155417970.jpg;Не удалось определить'],
    [
        '869874179;Acer 17.3"HD+ /nvidia GeForce GT 840m в Москве;27000;Не удалось определить;Сегодня18:31;https://www.avito.ru/moskva/noutbuki/acer_17.3hd_nvidia_geforce_gt_840m_869874179;//03.img.avito.st/140x105/3155416703.jpg;Не удалось определить'],
    [
        '308725113;Toshiba 305NB - 10E в Москве;8000;м.\xa0Медведково;Сегодня18:31;https://www.avito.ru/moskva/noutbuki/toshiba_305nb_-_10e_308725113;//60.img.avito.st/140x105/743921660.jpg;Не удалось определить'],
    [
        '869872749;Продам HP 15g-001sr в Москве;8000;м.\xa0Речной вокзал;Сегодня18:28;https://www.avito.ru/moskva/noutbuki/prodam_hp_15g-001sr_869872749;//31.img.avito.st/140x105/3155402531.jpg;Не удалось определить'],
    [
        '869383932;Macbook 12new grey 256Гб в Москве;69900;м.\xa0Планерная;Сегодня18:28;https://www.avito.ru/moskva/noutbuki/macbook_12new_grey_256gb_869383932;//36.img.avito.st/140x105/3153691136.jpg;Компания'],
    [
        '869872540;Огромный Dell Inspiron N7010 Core i5 в Москве;13000;м.\xa0Речной вокзал;Сегодня18:28;https://www.avito.ru/moskva/noutbuki/ogromnyy_dell_inspiron_n7010_core_i5_869872540;//04.img.avito.st/140x105/3155410304.jpg;Не удалось определить'],
    [
        '869871650;Новый игровой ноутбук MSI 6QF (GL626QF) в Москве;59000;м.\xa0Багратионовская;Сегодня18:26;https://www.avito.ru/moskva/noutbuki/novyy_igrovoy_noutbuk_msi_6qf_gl626qf_869871650;//22.img.avito.st/140x105/3155402122.jpg;Не удалось определить'],
    [
        '834340567;Ноутбук HP cn6320 на запчасти в Москве;2000;м.\xa0Новые Черемушки;Сегодня18:25;https://www.avito.ru/moskva/noutbuki/noutbuk_hp_cn6320_na_zapchasti_834340567;//95.img.avito.st/140x105/2966379395.jpg;Компания'],
    [
        '869869964;MSI X-Slim340 в Москве;10000;м.\xa0Тушинская;Сегодня18:23;https://www.avito.ru/moskva/noutbuki/msi_x-slim340_869869964;//29.img.avito.st/140x105/3155398429.jpg;Не удалось определить'],
    [
        '869869529;Ноутбук asus в Москве;10000;м.\xa0Орехово;Сегодня18:22;https://www.avito.ru/moskva/noutbuki/noutbuk_asus_869869529;//39.img.avito.st/140x105/3155397739.jpg;Компания'],
    [
        '869869410;MacBook Air 13" середина 2013года i5/HD5000 в Москве;35990;м.\xa0Курская;Сегодня18:22;https://www.avito.ru/moskva/noutbuki/macbook_air_13_seredina_2013goda_i5hd5000_869869410;//54.img.avito.st/140x105/3155397254.jpg;Компания'],
    [
        '831342109;Ноутбук Apple Maс Bоok Air 13 в Москве;46100;Компания;Сегодня18:21;https://www.avito.ru/moskva/noutbuki/noutbuk_apple_mas_book_air_13_831342109;//75.img.avito.st/140x105/2948854675.jpg;Не удалось определить'],
    [
        '869868505;Нет бук asus в Москве;10000;м.\xa0Царицыно;Сегодня18:20;https://www.avito.ru/moskva/noutbuki/net_buk_asus_869868505;Нет картинки;Не удалось определить'],
    [
        '817085474;Toshiba satellite a200-28n на запчасти в Москве;2000;м.\xa0Петровско-Разумовская;Сегодня18:20;https://www.avito.ru/moskva/noutbuki/toshiba_satellite_a200-28n_na_zapchasti_817085474;//68.img.avito.st/140x105/2867233068.jpg;Компания'],
    [
        '869868073;Классный мощный Acer в Москве;14990;м.\xa0Текстильщики;Сегодня18:19;https://www.avito.ru/moskva/noutbuki/klassnyy_moschnyy_acer_869868073;//13.img.avito.st/140x105/3155394813.jpg;Компания'],
    [
        '869866748;На запчасти самсунг в Москве;1000;м.\xa0Петровско-Разумовская;Сегодня18:17;https://www.avito.ru/moskva/noutbuki/na_zapchasti_samsung_869866748;Нет картинки;Компания'],
    [
        '869866672;MacBook Air 13 2015 i5/4gb/128ssd в Москве;54999;м.\xa0Кантемировская;Сегодня18:17;https://www.avito.ru/moskva/noutbuki/macbook_air_13_2015_i54gb128ssd_869866672;//98.img.avito.st/140x105/3155389798.jpg;Не удалось определить'],
    [
        '365323243;Fujitsu Siemens LifeBook C-1110D ноутбук в Москве;1200;м.\xa0Тимирязевская;Сегодня18:16;https://www.avito.ru/moskva/noutbuki/fujitsu_siemens_lifebook_c-1110d_noutbuk_365323243;//20.img.avito.st/140x105/915186420.jpg;Не удалось определить'],
    [
        '365323243;Fujitsu Siemens LifeBook C-1110D ноутбук в Москве;1200;м.\xa0Тимирязевская;Сегодня18:16;https://www.avito.ru/moskva/noutbuki/fujitsu_siemens_lifebook_c-1110d_noutbuk_365323243;//20.img.avito.st/140x105/915186420.jpg;Не удалось определить'],
    [
        '869866235;Apple MacBook Air 11 Early 2015 в Москве;14000;Не удалось определить;Сегодня18:16;https://www.avito.ru/moskva/noutbuki/apple_macbook_air_11_early_2015_869866235;//64.img.avito.st/140x105/3155388564.jpg;Не удалось определить'],
    [
        '869865984;Asus 15.6" GeForce GT в Москве;12500;м.\xa0Тимирязевская;Сегодня18:15;https://www.avito.ru/moskva/noutbuki/asus_15.6_geforce_gt_869865984;//29.img.avito.st/140x105/3155386529.jpg;Не удалось определить'],
    [
        '869864966;Нетбук asus 2 ядра в Москве;5000;м.\xa0Домодедовская;Сегодня18:14;https://www.avito.ru/moskva/noutbuki/netbuk_asus_2_yadra_869864966;//91.img.avito.st/140x105/3155383791.jpg;Не удалось определить'],
    [
        '869864080;Ноутбук acer 5100 целиком или на запчасти в Москве;4500;м.\xa0Планерная;Сегодня18:12;https://www.avito.ru/moskva/noutbuki/noutbuk_acer_5100_tselikom_ili_na_zapchasti_869864080;//49.img.avito.st/140x105/3155379949.jpg;Не удалось определить'],
    [
        '783470006;Peтина13 256 SSD 2014 в Москве;65200;Компания;Сегодня18:12;https://www.avito.ru/moskva/noutbuki/petina13_256_ssd_2014_783470006;//57.img.avito.st/140x105/2670457957.jpg;Не удалось определить'],
    [
        '869863859;Sony vaio японское качество в Москве;11500;м.\xa0Павелецкая;Сегодня18:12;https://www.avito.ru/moskva/noutbuki/sony_vaio_yaponskoe_kachestvo_869863859;//83.img.avito.st/140x105/3155379983.jpg;Не удалось определить'],
    [
        '869863398;Lenovo B570e в Москве;10000;м.\xa0Медведково;Сегодня18:11;https://www.avito.ru/moskva/noutbuki/lenovo_b570e_869863398;//92.img.avito.st/140x105/3155378092.jpg;Компания'],
    [
        '869862702;Sony vaio 15.6 в Москве;14500;м.\xa0Белорусская;Сегодня18:10;https://www.avito.ru/moskva/noutbuki/sony_vaio_15.6_869862702;//93.img.avito.st/140x105/3155376893.jpg;Не удалось определить'],
    [
        '851252267;Продаю ноутбук sony в Москве;11000;Компания;Сегодня18:10;https://www.avito.ru/moskva/noutbuki/prodayu_noutbuk_sony_851252267;//03.img.avito.st/140x105/3065470203.jpg;Не удалось определить'],
    [
        '852059112;Нетбук Samsung в Москве;5500;м.\xa0Царицыно;Сегодня18:02;https://www.avito.ru/moskva/noutbuki/netbuk_samsung_852059112;//18.img.avito.st/140x105/3070259318.jpg;Не удалось определить'],
    [
        '869858058;Шустрый Dell Inspiron (3шт. Usb, WiFi, Vga, WinXp) в Москве;3500;Компания;Сегодня18:02;https://www.avito.ru/moskva/noutbuki/shustryy_dell_inspiron_3sht._usb_wifi_vga_winxp_869858058;//59.img.avito.st/140x105/3155360859.jpg;Не удалось определить'],
    [
        '854764489;Мощный Lenovo ThinkPad x240 оптом с гарантией в Москве;19999;м.\xa0Новослободская;Сегодня18:02;https://www.avito.ru/moskva/noutbuki/moschnyy_lenovo_thinkpad_x240_optom_s_garantiey_854764489;//54.img.avito.st/140x105/3145160954.jpg;Не удалось определить'],
    [
        '822637668;Macbook air 13 i7 256 Gb в Москве;44000;м.\xa0Текстильщики;Сегодня17:59;https://www.avito.ru/moskva/noutbuki/macbook_air_13_i7_256_gb_822637668;//82.img.avito.st/140x105/2898612582.jpg;Компания'],
    [
        '869856420;Лучший лайфбук Футджитсу сименс (ноутбук) в Москве;6490;м.\xa0Беговая;Сегодня17:59;https://www.avito.ru/moskva/noutbuki/luchshiy_layfbuk_futdzhitsu_simens_noutbuk_869856420;//01.img.avito.st/140x105/3155355101.jpg;Не удалось определить'],
    [
        '869855444;MacBook 12 Retina 256SSD Комплекте Гарантия в Москве;65000;Не удалось определить;Сегодня17:57;https://www.avito.ru/moskva/noutbuki/macbook_12_retina_256ssd_komplekte_garantiya_869855444;//89.img.avito.st/140x105/3155329389.jpg;Не удалось определить'],
    [
        '869855031;Macbook Pro 15 Retina 2014 i7/8gb/256/gt650m в Москве;85000;м.\xa0Сходненская;Сегодня17:57;https://www.avito.ru/moskva/noutbuki/macbook_pro_15_retina_2014_i78gb256gt650m_869855031;//95.img.avito.st/140x105/3155348695.jpg;Не удалось определить'],
    [
        '837671303;Про13 2011 в Москве;28500;м.\xa0Кузьминки;Сегодня17:55;https://www.avito.ru/moskva/noutbuki/pro13_2011_837671303;//86.img.avito.st/140x105/2985850586.jpg;Компания'],
    [
        '731911309;MakcBook Po 13 i7 2.9GHz/8gb/HD 4000 в Москве;52990;Компания;Сегодня17:53;https://www.avito.ru/moskva/noutbuki/makcbook_po_13_i7_2.9ghz8gbhd_4000_731911309;//24.img.avito.st/140x105/2534851924.jpg;Не удалось определить'],
    [
        '869852987;Ноутбук для танков в Москве;11990;м.\xa0Новокосино;Сегодня17:53;https://www.avito.ru/moskva/noutbuki/noutbuk_dlya_tankov_869852987;//10.img.avito.st/140x105/3155340110.jpg;Не удалось определить'],
    [
        '706647146;Sony Sve1512 p950 2.1ghzx2/4gb/320gb/2gb hd3000 ид в Москве;17600;м.\xa0Планерная;Сегодня17:52;https://www.avito.ru/moskva/noutbuki/sony_sve1512_p950_2.1ghzx24gb320gb2gb_hd3000_id_706647146;//42.img.avito.st/140x105/2170999442.jpg;Компания'],
    [
        '607806726;Asus x550 Ультрабук 4-х яд core i3 3217/4gb/500gb в Москве;18000;м.\xa0Спартак;Сегодня17:52;https://www.avito.ru/moskva/noutbuki/asus_x550_ultrabuk_4-h_yad_core_i3_32174gb500gb_607806726;//86.img.avito.st/140x105/1795468786.jpg;Компания'],
    [
        '552832620;Asus K42J 14" игровой 4-х яд Core i5/4/320/Radeon в Москве;17500;м.\xa0Планерная;Сегодня17:52;https://www.avito.ru/moskva/noutbuki/asus_k42j_14_igrovoy_4-h_yad_core_i54320radeon_552832620;//73.img.avito.st/140x105/1588209073.jpg;Компания'],
    [
        '529322006;Packard bell 15.6" 4-x яд 2.3ghzx4/4gb/500gb/3gbGF в Москве;16000;м.\xa0Планерная;Сегодня17:51;https://www.avito.ru/moskva/noutbuki/packard_bell_15.6_4-x_yad_2.3ghzx44gb500gb3gbgf_529322006;//69.img.avito.st/140x105/1494686769.jpg;Компания'],
    [
        '834718745;Sony sve1512q1r/w Core i5 3210/4gb/500gb/7650radeo в Москве;20000;м.\xa0Планерная;Сегодня17:51;https://www.avito.ru/moskva/noutbuki/sony_sve1512q1rw_core_i5_32104gb500gb7650radeo_834718745;//78.img.avito.st/140x105/2968607778.jpg;Компания'],
    [
        '569942880;Acer 5750g с 3Gb Geforce 630, 4-x яд 2.3ghzx4/4gb в Москве;16990;м.\xa0Спартак;Сегодня17:50;https://www.avito.ru/moskva/noutbuki/acer_5750g_s_3gb_geforce_630_4-x_yad_2.3ghzx44gb_569942880;//78.img.avito.st/140x105/1660200578.jpg;Компания'],
    [
        '510607433;Asus x101Ch 4-x яд белый 1.6ghzx4/1gb озу/320/4час в Москве;7100;м.\xa0Киевская;Сегодня17:50;https://www.avito.ru/moskva/noutbuki/asus_x101ch_4-x_yad_belyy_1.6ghzx41gb_ozu3204chas_510607433;//09.img.avito.st/140x105/1419641609.jpg;Компания'],
    [
        '662724832;Неубиваемый ноутбук в Москве;13490;м.\xa0Курская;Сегодня17:50;https://www.avito.ru/moskva/noutbuki/neubivaemyy_noutbuk_662724832;//94.img.avito.st/140x105/3155339594.jpg;Компания'],
    [
        '565493859;Dell precision m4500 core i7, FHD, N-Vidia Quadro в Москве;22000;м.\xa0Перово;Сегодня17:50;https://www.avito.ru/moskva/noutbuki/dell_precision_m4500_core_i7_fhd_n-vidia_quadro_565493859;//84.img.avito.st/140x105/1643462184.jpg;Компания'],
    [
        '869851370;Супер Лептоп Эйсер в Москве;5990;м.\xa0Цветной бульвар;Сегодня17:50;https://www.avito.ru/moskva/noutbuki/super_leptop_eyser_869851370;//63.img.avito.st/140x105/3155336063.jpg;Не удалось определить'],
    [
        '869851122;Ноутбук Siemens Fujitsu Amilo в Москве;8000;м.\xa0Юго-Западная;Сегодня17:49;https://www.avito.ru/moskva/noutbuki/noutbuk_siemens_fujitsu_amilo_869851122;//30.img.avito.st/140x105/3155335530.jpg;Не удалось определить'],
    [
        '790297271;Нетбук XE700T1C в Москве;4990;м.\xa0Люблино;Сегодня17:48;https://www.avito.ru/moskva/noutbuki/netbuk_xe700t1c_790297271;//22.img.avito.st/140x105/2709690422.jpg;Компания'],
    [
        '643997247;Lenovo Helix core i5, SSD 128 в Москве;25000;м.\xa0Новогиреево;Сегодня17:48;https://www.avito.ru/moskva/noutbuki/lenovo_helix_core_i5_ssd_128_643997247;//27.img.avito.st/140x105/1926508127.jpg;Компания'],
    [
        '799999340;Core i7.12 дюймов, сенсорный dell e7240 в Москве;45000;м.\xa0Новогиреево;Сегодня17:45;https://www.avito.ru/moskva/noutbuki/core_i7.12_dyuymov_sensornyy_dell_e7240_799999340;//63.img.avito.st/140x105/2766319663.jpg;Компания'],
    [
        '828538905;Dell latitude e7450 core i7, nvidia Geforce 840m в Москве;70000;м.\xa0Новогиреево;Сегодня17:45;https://www.avito.ru/moskva/noutbuki/dell_latitude_e7450_core_i7_nvidia_geforce_840m_828538905;//47.img.avito.st/140x105/2932459447.jpg;Компания'],
    [
        '846670347;I5, 8gb, 256SSD, 3G модем Dell latitude e6330 в Москве;22000;м.\xa0Новогиреево;Сегодня17:44;https://www.avito.ru/moskva/noutbuki/i5_8gb_256ssd_3g_modem_dell_latitude_e6330_846670347;//13.img.avito.st/140x105/3038537513.jpg;Компания'],
    [
        '460186269;Irbis в Москве;1800;м.\xa0Первомайская;Сегодня17:44;https://www.avito.ru/moskva/noutbuki/irbis_460186269;//27.img.avito.st/140x105/1268108727.jpg;Не удалось определить'],
    [
        '797766864;Ультрабук Dell Latitude e7240 12 дюймов c core i5 в Москве;28000;м.\xa0Новогиреево;Сегодня17:44;https://www.avito.ru/moskva/noutbuki/ultrabuk_dell_latitude_e7240_12_dyuymov_c_core_i5_797766864;//25.img.avito.st/140x105/2753320525.jpg;Компания'],
    [
        '766302422;MSI GS70 Stealth в Москве;50000;м.\xa0Строгино;Сегодня17:44;https://www.avito.ru/moskva/noutbuki/msi_gs70_stealth_766302422;//00.img.avito.st/140x105/2569788900.jpg;Не удалось определить'],
    [
        '799991772;12 дюймов, core i5, SSD, 4GB, Dell latitude e6220 в Москве;15000;м.\xa0Новогиреево;Сегодня17:44;https://www.avito.ru/moskva/noutbuki/12_dyuymov_core_i5_ssd_4gb_dell_latitude_e6220_799991772;//43.img.avito.st/140x105/2766270443.jpg;Компания'],
    [
        '852022118;Игровой Asus X501U + мышка в комплекте в Москве;12900;м.\xa0Павелецкая;Сегодня17:44;https://www.avito.ru/moskva/noutbuki/igrovoy_asus_x501u_myshka_v_komplekte_852022118;//64.img.avito.st/140x105/3070021764.jpg;Магазин'],
    [
        '869847950;Apple MacBook Pro 15 with Retina display Mid 2014 в Москве;24000;Не удалось определить;Сегодня17:43;https://www.avito.ru/moskva/noutbuki/apple_macbook_pro_15_with_retina_display_mid_2014_869847950;//37.img.avito.st/140x105/3155323237.jpg;Не удалось определить'],
    [
        '851445855;Asus трансформер T200 в Москве;20000;м.\xa0Молодежная;Сегодня17:43;https://www.avito.ru/moskva/noutbuki/asus_transformer_t200_851445855;//33.img.avito.st/140x105/3066668533.jpg;Компания'],
    [
        '869847709;Sony vgn-fw11sr-me217n в Москве;15900;м.\xa0Волгоградский проспект;Сегодня17:43;https://www.avito.ru/moskva/noutbuki/sony_vgn-fw11sr-me217n_869847709;//71.img.avito.st/140x105/3155320171.jpg;None'],
    [
        '557918917;12.5 легкий core i7, Dell Latitude e6230 в Москве;25000;м.\xa0Новогиреево;Сегодня17:42;https://www.avito.ru/moskva/noutbuki/12.5_legkiy_core_i7_dell_latitude_e6230_557918917;//89.img.avito.st/140x105/1609770789.jpg;Компания'],
    [
        '722136521;Core i7, 13.3, металлический Dell latitude e6330 в Москве;22000;м.\xa0Новогиреево;Сегодня17:42;https://www.avito.ru/moskva/noutbuki/core_i7_13.3_metallicheskiy_dell_latitude_e6330_722136521;//08.img.avito.st/140x105/2227329508.jpg;Компания'],
    [
        '818310222;Нетбук Asus Eee PC 1005HA 2Gb 64Gb SSD или на 320 в Москве;6500;м.\xa0Отрадное;Сегодня17:41;https://www.avito.ru/moskva/noutbuki/netbuk_asus_eee_pc_1005ha_2gb_64gb_ssd_ili_na_320_818310222;//66.img.avito.st/140x105/2874082866.jpg;Не удалось определить'],
    [
        '651366323;Dell latitude e6330, Core i5, USB 3.0, 4GB в Москве;18000;м.\xa0Новогиреево;Сегодня17:41;https://www.avito.ru/moskva/noutbuki/dell_latitude_e6330_core_i5_usb_3.0_4gb_651366323;//53.img.avito.st/140x105/1953184253.jpg;Компания'],
    [
        '168455602;Ноутбук HP dv6-6b54er в Москве;;м.\xa0Коломенская;Сегодня17:41;https://www.avito.ru/moskva/noutbuki/noutbuk_hp_dv6-6b54er_168455602;//73.img.avito.st/140x105/358217973.jpg;Компания'],
    [
        '869846634;Игровой ноутбук i5(3.1Гц) /6gb/gt640-2gb/1000gb в Москве;24400;м.\xa0Тушинская;Сегодня17:41;https://www.avito.ru/moskva/noutbuki/igrovoy_noutbuk_i53.1gts_6gbgt640-2gb1000gb_869846634;//52.img.avito.st/140x105/3155313752.jpg;Не удалось определить'],
    [
        '571404200;12.5 дюймов core i7, 4GB, SSD-128, Dell e6220 в Москве;22000;м.\xa0Новогиреево;Сегодня17:41;https://www.avito.ru/moskva/noutbuki/12.5_dyuymov_core_i7_4gb_ssd-128_dell_e6220_571404200;//59.img.avito.st/140x105/2305096659.jpg;Компания'],
    [
        '758097354;Lenovo Thinkpad T420 nVidia quadro + SSD в Москве;33490;м.\xa0Савеловская;Сегодня17:41;https://www.avito.ru/moskva/noutbuki/lenovo_thinkpad_t420_nvidia_quadro_ssd_758097354;//36.img.avito.st/140x105/3155345536.jpg;Компания'],
    [
        '869846068;Toshiba P100-RW215Q37 в Москве;10500;м.\xa0Волгоградский проспект;Сегодня17:40;https://www.avito.ru/moskva/noutbuki/toshiba_p100-rw215q37_869846068;//00.img.avito.st/140x105/3155314000.jpg;None'],
    [
        '849367751;Asus N75SF Intel Core i5 + сабвуфер в Москве;27500;м.\xa0Петровско-Разумовская;Сегодня17:39;https://www.avito.ru/moskva/noutbuki/asus_n75sf_intel_core_i5_sabvufer_849367751;//80.img.avito.st/140x105/3054672480.jpg;Не удалось определить'],
    [
        '852043502;Ноутбук Acer V3 771G по запчастям в Москве;300;м.\xa0Беляево;Сегодня17:38;https://www.avito.ru/moskva/noutbuki/noutbuk_acer_v3_771g_po_zapchastyam_852043502;//65.img.avito.st/140x105/3070149665.jpg;Компания'],
    [
        '869845038;MacBook Pro 2011 Идеальный i7/8gb/500gb в Москве;45000;Не удалось определить;Сегодня17:38;https://www.avito.ru/moskva/noutbuki/macbook_pro_2011_idealnyy_i78gb500gb_869845038;//22.img.avito.st/140x105/3155310022.jpg;Не удалось определить'],
    [
        '869844828;Sony vaio Pro 13 сверхлегкий ультрабук в Москве;41000;м.\xa0Пятницкое шоссе;Сегодня17:37;https://www.avito.ru/moskva/noutbuki/sony_vaio_pro_13_sverhlegkiy_ultrabuk_869844828;//25.img.avito.st/140x105/3155301825.jpg;Компания'],
    [
        '869844825;Acer 5920G-W25Q3712 в Москве;11600;м.\xa0Волгоградский проспект;Сегодня17:37;https://www.avito.ru/moskva/noutbuki/acer_5920g-w25q3712_869844825;//21.img.avito.st/140x105/3155308121.jpg;None'],
    [
        '869844263;Ноутбук+ портативный принтер в Москве;3000;м.\xa0Ясенево;Сегодня17:36;https://www.avito.ru/moskva/noutbuki/noutbuk_portativnyy_printer_869844263;//14.img.avito.st/140x105/3155307314.jpg;Не удалось определить'],
    [
        '869843838;Acer 5520 в Москве;3000;м.\xa0Первомайская;Сегодня17:35;https://www.avito.ru/moskva/noutbuki/acer_5520_869843838;//39.img.avito.st/140x105/3155306339.jpg;Компания'],
    [
        '747570361;Core i5.4 ядра металлический Dell e6420 в Москве;14000;м.\xa0Новогиреево;Сегодня17:34;https://www.avito.ru/moskva/noutbuki/core_i5.4_yadra_metallicheskiy_dell_e6420_747570361;//16.img.avito.st/140x105/2417251916.jpg;Компания'],
    [
        '869843257;Retina 12 в Москве;60000;м.\xa0Рижская;Сегодня17:34;https://www.avito.ru/moskva/noutbuki/retina_12_869843257;Нет картинки;Компания'],
    [
        '747623190;14 дюймов компактный dell e7440 в Москве;28000;м.\xa0Новогиреево;Сегодня17:34;https://www.avito.ru/moskva/noutbuki/14_dyuymov_kompaktnyy_dell_e7440_747623190;//16.img.avito.st/140x105/2417703316.jpg;Компания'],
    [
        '279512960;Asus 1010b мощный игровой Amd c30/2/320/1gb radeon в Москве;7490;Компания;Сегодня17:34;https://www.avito.ru/moskva/noutbuki/asus_1010b_moschnyy_igrovoy_amd_c3023201gb_radeon_279512960;//14.img.avito.st/140x105/2244543214.jpg;Не удалось определить'],
    [
        '648167770;Надежный, квадратный в хорошем состояни Lenovo T60 в Москве;7990;м.\xa0Савеловская;Сегодня17:34;https://www.avito.ru/moskva/noutbuki/nadezhnyy_kvadratnyy_v_horoshem_sostoyani_lenovo_t60_648167770;//26.img.avito.st/140x105/1941500226.jpg;Компания'],
    [
        '747612374;Ультра бук core i7 Dell Latitude e7240 в Москве;38000;м.\xa0Новогиреево;Сегодня17:34;https://www.avito.ru/moskva/noutbuki/ultra_buk_core_i7_dell_latitude_e7240_747612374;//44.img.avito.st/140x105/2462529844.jpg;Компания'],
    [
        '615257963;Dell xps 11, core i5, сенсорный трансфмер 2 в 1 в Москве;30000;м.\xa0Новогиреево;Сегодня17:33;https://www.avito.ru/moskva/noutbuki/dell_xps_11_core_i5_sensornyy_transfmer_2_v_1_615257963;//69.img.avito.st/140x105/1820838869.jpg;Компания'],
    [
        '655640471;Dell precision m6500 core i5, 4GB в Москве;25000;м.\xa0Новогиреево;Сегодня17:33;https://www.avito.ru/moskva/noutbuki/dell_precision_m6500_core_i5_4gb_655640471;//47.img.avito.st/140x105/2417629547.jpg;Компания'],
    [
        '524493188;12 дюйм core i5, USB 3.0.4GB, Dell latitude e6230 в Москве;18000;м.\xa0Новогиреево;Сегодня17:32;https://www.avito.ru/moskva/noutbuki/12_dyuym_core_i5_usb_3.0.4gb_dell_latitude_e6230_524493188;//82.img.avito.st/140x105/2305107582.jpg;Компания'],
    [
        '551944622;15.6 core i5.4 гига, 320 GB Dell e6510 HD в Москве;14000;м.\xa0Новогиреево;Сегодня17:32;https://www.avito.ru/moskva/noutbuki/15.6_core_i5.4_giga_320_gb_dell_e6510_hd_551944622;//48.img.avito.st/140x105/1584637048.jpg;Компания'],
    [
        '808242357;Dell precision 5510, ultra HD, процессор Xeon в Москве;135000;м.\xa0Новогиреево;Сегодня17:32;https://www.avito.ru/moskva/noutbuki/dell_precision_5510_ultra_hd_protsessor_xeon_808242357;//89.img.avito.st/140x105/2963844289.jpg;Компания'],
    [
        '869842022;Мощный, 8 ядерный Dell. (Core i7-2630qm) в Москве;19000;Не удалось определить;Сегодня17:32;https://www.avito.ru/moskva/noutbuki/moschnyy_8_yadernyy_dell._core_i7-2630qm_869842022;//35.img.avito.st/140x105/3155297535.jpg;Не удалось определить'],
    [
        '869841998;Fujitsu Siemens MS2239-N25R-MT217 в Москве;11200;м.\xa0Волгоградский проспект;Сегодня17:32;https://www.avito.ru/moskva/noutbuki/fujitsu_siemens_ms2239-n25r-mt217_869841998;//34.img.avito.st/140x105/3155297534.jpg;None'],
    [
        '869841966;Новый игровой ноутбук dell alienware 13 R2 P56G002 в Москве;57000;м.\xa0Багратионовская;Сегодня17:31;https://www.avito.ru/moskva/noutbuki/novyy_igrovoy_noutbuk_dell_alienware_13_r2_p56g002_869841966;//92.img.avito.st/140x105/3155288992.jpg;Не удалось определить'],
    [
        '567923076;13.3, core i5, 4GB, SSD, Dell Latitude e6320 в Москве;16000;м.\xa0Новогиреево;Сегодня17:31;https://www.avito.ru/moskva/noutbuki/13.3_core_i5_4gb_ssd_dell_latitude_e6320_567923076;//03.img.avito.st/140x105/2394048803.jpg;Компания'],
    [
        '689395591;Dell xps 12, сенсорный ультрабук трансформер с i7 в Москве;35000;м.\xa0Новогиреево;Сегодня17:31;https://www.avito.ru/moskva/noutbuki/dell_xps_12_sensornyy_ultrabuk_transformer_s_i7_689395591;//53.img.avito.st/140x105/2100707653.jpg;Компания'],
    [
        '869841511;Пластиковый чехол MacBook Pro 13 Салатовый в Москве;500;м.\xa0Выхино;Сегодня17:31;https://www.avito.ru/moskva/noutbuki/plastikovyy_chehol_macbook_pro_13_salatovyy_869841511;//70.img.avito.st/140x105/3155297070.jpg;Не удалось определить'],
    [
        '869840665;Macbook air a1237 в Москве;8000;м.\xa0Арбатская;Сегодня17:29;https://www.avito.ru/moskva/noutbuki/macbook_air_a1237_869840665;//86.img.avito.st/140x105/3155292586.jpg;Компания'],
    [
        '869839742;Samsung R410-ME216Q-5725 в Москве;12900;м.\xa0Волгоградский проспект;Сегодня17:27;https://www.avito.ru/moskva/noutbuki/samsung_r410-me216q-5725_869839742;//66.img.avito.st/140x105/3155289166.jpg;None'],
    [
        '869839722;Ноутбук Acer TJ75 (i5, 8gb, 500gb, 5850 1gb) в Москве;14900;Не удалось определить;Сегодня17:27;https://www.avito.ru/moskva/noutbuki/noutbuk_acer_tj75_i5_8gb_500gb_5850_1gb_869839722;//03.img.avito.st/140x105/3155289403.jpg;Не удалось определить'],
    [
        '848701753;Ноутбук HP 250 1g в Москве;11000;м.\xa0Бабушкинская;Сегодня17:27;https://www.avito.ru/moskva/noutbuki/noutbuk_hp_250_1g_848701753;//89.img.avito.st/140x105/3050745689.jpg;Компания'],
    [
        '869839108;Sony Vaio 15.6" в Москве;12500;м.\xa0Новокосино;Сегодня17:26;https://www.avito.ru/moskva/noutbuki/sony_vaio_15.6_869839108;//16.img.avito.st/140x105/3155281816.jpg;Не удалось определить'],
    [
        '852033893;Защищенный ноутбук ленoво Сore i5 в Москве;22200;м.\xa0Кантемировская;Сегодня17:25;https://www.avito.ru/moskva/noutbuki/zaschischennyy_noutbuk_lenovo_sore_i5_852033893;//06.img.avito.st/140x105/3070088206.jpg;Компания'],
    [
        '869838470;HP 635-N24WCR57Q2 в Москве;15900;м.\xa0Волгоградский проспект;Сегодня17:24;https://www.avito.ru/moskva/noutbuki/hp_635-n24wcr57q2_869838470;//59.img.avito.st/140x105/3155283559.jpg;None'],
    [
        '869838301;Нетбук lenovo u165 в Москве;3000;м.\xa0Отрадное;Сегодня17:24;https://www.avito.ru/moskva/noutbuki/netbuk_lenovo_u165_869838301;//32.img.avito.st/140x105/3155282732.jpg;Не удалось определить'],
    [
        '869837271;Нетбука HP Mini 210-2210er в Москве;6000;м.\xa0Теплый стан;Сегодня17:22;https://www.avito.ru/moskva/noutbuki/netbuka_hp_mini_210-2210er_869837271;//86.img.avito.st/140x105/3155219086.jpg;Не удалось определить'],
    [
        '869837082;Нр новый игровой в Москве;15000;м.\xa0Алтуфьево;Сегодня17:22;https://www.avito.ru/moskva/noutbuki/nr_novyy_igrovoy_869837082;//84.img.avito.st/140x105/3155279484.jpg;Не удалось определить'],
    [
        '869837049;MacBook Pro 13 2012 Идеальный в Москве;28000;Не удалось определить;Сегодня17:22;https://www.avito.ru/moskva/noutbuki/macbook_pro_13_2012_idealnyy_869837049;//39.img.avito.st/140x105/3155274339.jpg;Не удалось определить'],
    [
        '869836985;Ноутбук Asus в Москве;12970;м.\xa0Киевская;Сегодня17:22;https://www.avito.ru/moskva/noutbuki/noutbuk_asus_869836985;//82.img.avito.st/140x105/3155276582.jpg;Компания'],
    [
        '869836848;Dell inspiron 1300-me215 в Москве;7400;м.\xa0Волгоградский проспект;Сегодня17:21;https://www.avito.ru/moskva/noutbuki/dell_inspiron_1300-me215_869836848;//17.img.avito.st/140x105/3155277017.jpg;None'],
    [
        '869836249;Ноутбук для офиса в Москве;15500;м.\xa0Новокосино;Сегодня17:20;https://www.avito.ru/moskva/noutbuki/noutbuk_dlya_ofisa_869836249;//61.img.avito.st/140x105/3155273461.jpg;Не удалось определить'],
    [
        '869835978;Нужняя крышка корпуса HP 6820S в Москве;1000;м.\xa0Лермонтовский проспект;Сегодня17:20;https://www.avito.ru/moskva/noutbuki/nuzhnyaya_kryshka_korpusa_hp_6820s_869835978;//57.img.avito.st/140x105/3155270057.jpg;Не удалось определить'],
    [
        '783211960;Компактный и надежный Dell в отл. состоянии в Москве;9990;м.\xa0Савеловская;Сегодня17:19;https://www.avito.ru/moskva/noutbuki/kompaktnyy_i_nadezhnyy_dell_v_otl._sostoyanii_783211960;//96.img.avito.st/140x105/2668941396.jpg;Компания'],
    [
        '847780871;В Идеале Восьми-Ядерный hp i7 с Батареей 4ч, 3-Gb в Москве;16500;Компания;Сегодня17:18;https://www.avito.ru/moskva/noutbuki/v_ideale_vosmi-yadernyy_hp_i7_s_batareey_4ch_3-gb_847780871;//17.img.avito.st/140x105/3045201917.jpg;Не удалось определить'],
    [
        '849892007;Нужен как "воздух" в Москве;38600;м.\xa0Крестьянская застава;Сегодня17:17;https://www.avito.ru/moskva/noutbuki/nuzhen_kak_vozduh_849892007;//96.img.avito.st/140x105/3057687596.jpg;Компания'],
    [
        '869834366;HP ProBook 14" /Core i5 - 2.70ghz/500gb в Москве;23000;Не удалось определить;Сегодня17:16;https://www.avito.ru/moskva/noutbuki/hp_probook_14_core_i5_-_2.70ghz500gb_869834366;//92.img.avito.st/140x105/3155266192.jpg;Не удалось определить'],
    [
        '456873983;15" Тошиба с живым 2 часовым Аккум 1500mhz/1gb/dvd в Москве;5880;м.\xa0Охотный ряд;Сегодня17:16;https://www.avito.ru/moskva/noutbuki/15_toshiba_s_zhivym_2_chasovym_akkum_1500mhz1gbdvd_456873983;//47.img.avito.st/140x105/1209750347.jpg;Компания'],
    [
        '869834080;Dell XPS 12 новый в Москве;77500;м.\xa0Багратионовская;Сегодня17:16;https://www.avito.ru/moskva/noutbuki/dell_xps_12_novyy_869834080;//93.img.avito.st/140x105/3155264193.jpg;Не удалось определить'],
    [
        '384229990;Крутой Acer 5000 4-х яд. Core i5 2.67x4/8gb/500/GF в Москве;17990;м.\xa0Планерная;Сегодня17:15;https://www.avito.ru/moskva/noutbuki/krutoy_acer_5000_4-h_yad._core_i5_2.67x48gb500gf_384229990;//86.img.avito.st/140x105/979028686.jpg;Компания'],
    [
        '849910396;Sony vaio VPC-F13Z8R в Москве;25000;м.\xa0Краснопресненская;Сегодня17:15;https://www.avito.ru/moskva/noutbuki/sony_vaio_vpc-f13z8r_849910396;//71.img.avito.st/140x105/3057794671.jpg;Не удалось определить'],
    [
        '783250270;IBM - Lenovo thinkpad T61 в Москве;8990;м.\xa0Савеловская;Сегодня17:14;https://www.avito.ru/moskva/noutbuki/ibm_-_lenovo_thinkpad_t61_783250270;//91.img.avito.st/140x105/2669161291.jpg;Компания'],
    [
        '869833140;Игровой MSI (i7 + 8 Гб озу+ geforce GTX 750 ) в Москве;41990;м.\xa0Арбатская;Сегодня17:14;https://www.avito.ru/moskva/noutbuki/igrovoy_msi_i7_8_gb_ozu_geforce_gtx_750_869833140;//32.img.avito.st/140x105/3155263132.jpg;Не удалось определить'],
    [
        '869832909;Сверх-надёжный acer BL52 в Москве;6000;м.\xa0Римская;Сегодня17:13;https://www.avito.ru/moskva/noutbuki/sverh-nadezhnyy_acer_bl52_869832909;//44.img.avito.st/140x105/3155259544.jpg;Не удалось определить'],
    [
        '869832096;Новый MSI GT72S 6QE dominator PRO G tobii в Москве;155500;м.\xa0Фили;Сегодня17:12;https://www.avito.ru/moskva/noutbuki/novyy_msi_gt72s_6qe_dominator_pro_g_tobii_869832096;//73.img.avito.st/140x105/3155252573.jpg;Компания'],
    [
        '861689927;Ноутбук Acer Aspire TimelineX 5820TG - Core i5 4GB в Москве;12800;м.\xa0Щелковская;Сегодня17:11;https://www.avito.ru/moskva/noutbuki/noutbuk_acer_aspire_timelinex_5820tg_-_core_i5_4gb_861689927;//93.img.avito.st/140x105/3125684093.jpg;Не удалось определить'],
    [
        '869831609;MacBook Pro 13 /4Gb/500Gb в Москве;22000;м.\xa0Владыкино;Сегодня17:11;https://www.avito.ru/moskva/noutbuki/macbook_pro_13_4gb500gb_869831609;//07.img.avito.st/140x105/3155299307.jpg;Не удалось определить'],
    [
        '869831549;Hp ProBook 15.6 в Москве;32000;м.\xa0Лермонтовский проспект;Сегодня17:11;https://www.avito.ru/moskva/noutbuki/hp_probook_15.6_869831549;//28.img.avito.st/140x105/3155279528.jpg;Не удалось определить'],
    [
        '868840889;MacBook Pro (Retina, 15-inch, Mid 2015) в Москве;110000;м.\xa0Преображенская площадь;Сегодня17:11;https://www.avito.ru/moskva/noutbuki/macbook_pro_retina_15-inch_mid_2015_868840889;//57.img.avito.st/140x105/3151737957.jpg;Компания'],
    [
        '869830767;IBM t42 неубиваемый в Москве;3500;м.\xa0Добрынинская;Сегодня17:09;https://www.avito.ru/moskva/noutbuki/ibm_t42_neubivaemyy_869830767;//22.img.avito.st/140x105/3155259022.jpg;Компания'],
    [
        '829363204;MacBook 13 в Москве;22500;Компания;Сегодня17:09;https://www.avito.ru/moskva/noutbuki/macbook_13_829363204;//36.img.avito.st/140x105/2937331836.jpg;Не удалось определить'],
    [
        '819650189;Alienware 18 в Москве;115000;м.\xa0Молодежная;Сегодня17:08;https://www.avito.ru/moskva/noutbuki/alienware_18_819650189;//92.img.avito.st/140x105/2881872892.jpg;Не удалось определить'],
    [
        '851932713;Макбукпро 15 2012г i7/8gb/750/nVidia 650M/62 цикла в Москве;49500;м.\xa0Кузьминки;Сегодня17:07;https://www.avito.ru/moskva/noutbuki/makbukpro_15_2012g_i78gb750nvidia_650m62_tsikla_851932713;Нет картинки;Компания'],
    [
        '869829045;MacBook Pro 13 дюймов Retina display Mid 2015 в Москве;23000;Не удалось определить;Сегодня17:06;https://www.avito.ru/moskva/noutbuki/macbook_pro_13_dyuymov_retina_display_mid_2015_869829045;//93.img.avito.st/140x105/3155246693.jpg;Не удалось определить'],
    [
        '844698143;Продаю ноутбук fijutsu siemens amilo Xa1526 в Москве;6000;м.\xa0Новые Черемушки;Сегодня17:06;https://www.avito.ru/moskva/noutbuki/prodayu_noutbuk_fijutsu_siemens_amilo_xa1526_844698143;//72.img.avito.st/140x105/3026938972.jpg;Не удалось определить'],
    [
        '869828743;MacBook Pro 15 2013 Retina i7/8Gb/GT650M/SSD512gb в Москве;85000;Не удалось определить;Сегодня17:05;https://www.avito.ru/moskva/noutbuki/macbook_pro_15_2013_retina_i78gbgt650mssd512gb_869828743;//82.img.avito.st/140x105/3155215782.jpg;Не удалось определить'],
    [
        '101999379;Asus VX2S-Lamborghin в Москве;13500;м.\xa0Новогиреево;Сегодня17:05;https://www.avito.ru/moskva/noutbuki/asus_vx2s-lamborghin_101999379;//06.img.avito.st/140x105/190920406.jpg;Не удалось определить'],
    [
        '730788575;Macbook Air 13 MD760RU/А в Москве;41990;м.\xa0Багратионовская;Сегодня17:05;https://www.avito.ru/moskva/noutbuki/macbook_air_13_md760rua_730788575;//24.img.avito.st/140x105/3155243024.jpg;Магазин'],
    [
        '869827834;Новый MSI GT72S 6QF dominator PRO G в Москве;191000;м.\xa0Фили;Сегодня17:03;https://www.avito.ru/moskva/noutbuki/novyy_msi_gt72s_6qf_dominator_pro_g_869827834;//88.img.avito.st/140x105/3155236288.jpg;Компания'],
    [
        '869827662;Dell XPS 13 в Москве;45000;м.\xa0Войковская;Сегодня17:03;https://www.avito.ru/moskva/noutbuki/dell_xps_13_869827662;//54.img.avito.st/140x105/3155241054.jpg;Не удалось определить'],
    [
        '835477557;Ноутбук asus F83VD (F83V) в Москве;6000;Не удалось определить;Сегодня17:03;https://www.avito.ru/moskva/noutbuki/noutbuk_asus_f83vd_f83v_835477557;//06.img.avito.st/140x105/2973082606.jpg;Не удалось определить'],
    [
        '314301653;Супер мощный 15" HP Pavilion 6.3ghz/4096/490gb/2g в Москве;13000;м.\xa0Планерная;Сегодня17:02;https://www.avito.ru/moskva/noutbuki/super_moschnyy_15_hp_pavilion_6.3ghz4096490gb2g_314301653;//79.img.avito.st/140x105/759341979.jpg;Компания'],
    [
        '869826467;Apple MacBook 11 дюймов 2012 года в Москве;40000;м.\xa0Митино;Сегодня17:01;https://www.avito.ru/moskva/noutbuki/apple_macbook_11_dyuymov_2012_goda_869826467;//16.img.avito.st/140x105/3155235616.jpg;Компания'],
    [
        '869826207;MacBook Pro 15 2011 запчасти в Москве;1000;м.\xa0Тушинская;Сегодня17:00;https://www.avito.ru/moskva/noutbuki/macbook_pro_15_2011_zapchasti_869826207;//31.img.avito.st/140x105/3155232931.jpg;Не удалось определить'],
    [
        '291875600;15" apple macbook pro алюминий в Москве;22000;Компания;Сегодня16:59;https://www.avito.ru/moskva/noutbuki/15_apple_macbook_pro_alyuminiy_291875600;//24.img.avito.st/140x105/698108724.jpg;Не удалось определить'],
    [
        '869825052;Most balanced, HP ultrabook на Core i7 в Москве;25990;м.\xa0Марьина Роща;Сегодня16:58;https://www.avito.ru/moskva/noutbuki/most_balanced_hp_ultrabook_na_core_i7_869825052;//08.img.avito.st/140x105/3155229108.jpg;Компания'],
    [
        '852017815;Как новый бизнес ноутбук i5 8gb ram 14" в Москве;22000;м.\xa0Каширская;Сегодня16:58;https://www.avito.ru/moskva/noutbuki/kak_novyy_biznes_noutbuk_i5_8gb_ram_14_852017815;//91.img.avito.st/140x105/3078168691.jpg;Компания'],
    [
        '869824592;Asus M60V в Москве;15000;м.\xa0Кузьминки;Сегодня16:57;https://www.avito.ru/moskva/noutbuki/asus_m60v_869824592;//22.img.avito.st/140x105/3155228422.jpg;Не удалось определить'],
    [
        '769246354;Продам игровой ноутбук-ультрабук asus K56CM в Москве;25499;м.\xa0Братиславская;Сегодня16:57;https://www.avito.ru/moskva/noutbuki/prodam_igrovoy_noutbuk-ultrabuk_asus_k56cm_769246354;//98.img.avito.st/140x105/2588727698.jpg;Компания'],
    [
        '587454690;Ноутбук asus zenbook UX 305 FA в Москве;62999;Не удалось определить;Сегодня16:57;https://www.avito.ru/moskva/noutbuki/noutbuk_asus_zenbook_ux_305_fa_587454690;//96.img.avito.st/140x105/1723968996.jpg;Не удалось определить'],
    [
        '869823913;Macbook Pro 15 2012 Nvidia GT650M, DDR 8Гб, i7 2.6 в Москве;56000;м.\xa0Юго-Западная;Сегодня16:56;https://www.avito.ru/moskva/noutbuki/macbook_pro_15_2012_nvidia_gt650m_ddr_8gb_i7_2.6_869823913;//84.img.avito.st/140x105/3155224584.jpg;Не удалось определить'],
    [
        '833984322;Ноутбуки по запчастям в Москве;1000;м.\xa0Войковская;Сегодня16:54;https://www.avito.ru/moskva/noutbuki/noutbuki_po_zapchastyam_833984322;Нет картинки;Компания'],
    [
        '869823227;Мощный, надежный ноутбук Sony Vaio в Москве;8500;Не удалось определить;Сегодня16:54;https://www.avito.ru/moskva/noutbuki/moschnyy_nadezhnyy_noutbuk_sony_vaio_869823227;//41.img.avito.st/140x105/3155219841.jpg;Не удалось определить'],
    [
        '869823182;Asus 15.4 Turion 64 2.0Ghz 2GB HDD 320 Зеленоград в Москве;5500;м.\xa0Речной вокзал;Сегодня16:54;https://www.avito.ru/moskva/noutbuki/asus_15.4_turion_64_2.0ghz_2gb_hdd_320_zelenograd_869823182;//30.img.avito.st/140x105/3155220630.jpg;Не удалось определить'],
    [
        '869822417;Apple Macbook pro Core i5 SSD в Москве;38000;м.\xa0Щелковская;Сегодня16:53;https://www.avito.ru/moskva/noutbuki/apple_macbook_pro_core_i5_ssd_869822417;//27.img.avito.st/140x105/3155157127.jpg;Не удалось определить'],
    [
        '869822312;Новый asus ROG G752VY-DH72 в Москве;133000;м.\xa0Фили;Сегодня16:53;https://www.avito.ru/moskva/noutbuki/novyy_asus_rog_g752vy-dh72_869822312;//32.img.avito.st/140x105/3155213232.jpg;Компания'],
    [
        '751689488;Продам по запчастям Asus X50Z в Москве;3000;м.\xa0Крылатское;Сегодня16:50;https://www.avito.ru/moskva/noutbuki/prodam_po_zapchastyam_asus_x50z_751689488;//29.img.avito.st/140x105/2450701229.jpg;Не удалось определить'],
    [
        '868086000;Современный ультрабук Dell E6230 -i7, 6Gb, 250 ssd в Москве;19000;м.\xa0Алексеевская;Сегодня16:50;https://www.avito.ru/moskva/noutbuki/sovremennyy_ultrabuk_dell_e6230_-i7_6gb_250_ssd_868086000;//01.img.avito.st/140x105/3150868301.jpg;Не удалось определить'],
    [
        '869819908;Клавиатура для Asus Eee PC 1015 белая б/у в Москве;1000;м.\xa0Щелковская;Сегодня16:48;https://www.avito.ru/moskva/noutbuki/klaviatura_dlya_asus_eee_pc_1015_belaya_bu_869819908;Нет картинки;Не удалось определить'],
    [
        '654491175;Тактические, надеждные, легендарные IBM-Lenovo в Москве;12990;м.\xa0Савеловская;Сегодня16:46;https://www.avito.ru/moskva/noutbuki/takticheskie_nadezhdnye_legendarnye_ibm-lenovo_654491175;//82.img.avito.st/140x105/2290959882.jpg;Компания'],
    [
        '869818648;Sony 15.6 FHD в Москве;20000;м.\xa0Октябрьское поле;Сегодня16:46;https://www.avito.ru/moskva/noutbuki/sony_15.6_fhd_869818648;//22.img.avito.st/140x105/3155204022.jpg;Не удалось определить'],
    [
        '852006870;Сверхлёгкий Стильный i7 HasweI, SSD в Москве;38400;м.\xa0Каширская;Сегодня16:46;https://www.avito.ru/moskva/noutbuki/sverhlegkiy_stilnyy_i7_haswei_ssd_852006870;//49.img.avito.st/140x105/3069923449.jpg;Компания'],
    [
        '863999217;Продаю ноутбук Sony Vaio SVT1312M1RS в Москве;15000;м.\xa0Сходненская;Сегодня16:45;https://www.avito.ru/moskva/noutbuki/prodayu_noutbuk_sony_vaio_svt1312m1rs_863999217;//39.img.avito.st/140x105/3134190439.jpg;Не удалось определить'],
    [
        '869818197;Macbook Air (13-inch, early 2014) 128 SSD в Москве;42000;м.\xa0Белорусская;Сегодня16:45;https://www.avito.ru/moskva/noutbuki/macbook_air_13-inch_early_2014_128_ssd_869818197;//17.img.avito.st/140x105/3155201617.jpg;Не удалось определить'],
    [
        '634183836;Asus Zenbook UX32A на запчасти в Москве;7000;м.\xa0Волжская;Сегодня16:42;https://www.avito.ru/moskva/noutbuki/asus_zenbook_ux32a_na_zapchasti_634183836;//59.img.avito.st/140x105/3155224059.jpg;Не удалось определить'],
    [
        '869816144;Система охлаждения ноутбука HP P/N 643363-001 в Москве;450;м.\xa0Щелковская;Сегодня16:41;https://www.avito.ru/moskva/noutbuki/sistema_ohlazhdeniya_noutbuka_hp_pn_643363-001_869816144;Нет картинки;Не удалось определить'],
    [
        '869815942;MacBook Pro retina 15 mid 2015 в Москве;99000;м.\xa0Бабушкинская;Сегодня16:41;https://www.avito.ru/moskva/noutbuki/macbook_pro_retina_15_mid_2015_869815942;//55.img.avito.st/140x105/3155200755.jpg;Не удалось определить'],
    [
        '671357695;Игровой Lenovo 4-ядра, GeForce - 2GB с гарантией в Москве;12000;Компания;Сегодня16:40;https://www.avito.ru/moskva/noutbuki/igrovoy_lenovo_4-yadra_geforce_-_2gb_s_garantiey_671357695;//40.img.avito.st/140x105/2030133440.jpg;Не удалось определить'],
    [
        '844055318;Asus Eee PC-1008P в Москве;8000;м.\xa0Октябрьское поле;Сегодня16:38;https://www.avito.ru/moskva/noutbuki/asus_eee_pc-1008p_844055318;//46.img.avito.st/140x105/3023084546.jpg;Не удалось определить'],
    [
        '574370779;Fujitsu siemens lifebook в Москве;4000;м.\xa0Медведково;Сегодня16:38;https://www.avito.ru/moskva/noutbuki/fujitsu_siemens_lifebook_574370779;//27.img.avito.st/140x105/1676200727.jpg;Не удалось определить'],
    [
        '795153327;Asus PRO57T на запчасти в Москве;2000;м.\xa0Отрадное;Сегодня16:38;https://www.avito.ru/moskva/noutbuki/asus_pro57t_na_zapchasti_795153327;//36.img.avito.st/140x105/2738038236.jpg;Компания'],
    [
        '869814229;Ноутбук acer Aspire V в Москве;24999;м.\xa0Добрынинская;Сегодня16:38;https://www.avito.ru/moskva/noutbuki/noutbuk_acer_aspire_v_869814229;//87.img.avito.st/140x105/3155181987.jpg;Не удалось определить'],
    [
        '511317121;Acer E5-571g, новый в коробке i5 Haswell, 3gb GF82 в Москве;23980;м.\xa0Планерная;Сегодня16:37;https://www.avito.ru/moskva/noutbuki/acer_e5-571g_novyy_v_korobke_i5_haswell_3gb_gf82_511317121;//10.img.avito.st/140x105/1422234610.jpg;Компания'],
    [
        '869813031;Apple Macbook Air 11 i5 ssd128Гб в Москве;32000;м.\xa0Аэропорт;Сегодня16:36;https://www.avito.ru/moskva/noutbuki/apple_macbook_air_11_i5_ssd128gb_869813031;//49.img.avito.st/140x105/3155177049.jpg;Не удалось определить'],
    [
        '869812953;Мощный Sony Vaio 14 Core i3 4GB HDD 320 GF330 в Москве;10000;м.\xa0Волоколамская;Сегодня16:36;https://www.avito.ru/moskva/noutbuki/moschnyy_sony_vaio_14_core_i3_4gb_hdd_320_gf330_869812953;//51.img.avito.st/140x105/3155188551.jpg;Не удалось определить'],
    [
        '869812824;HP 255 G3 в Москве;8300;м.\xa0Петровско-Разумовская;Сегодня16:36;https://www.avito.ru/moskva/noutbuki/hp_255_g3_869812824;//17.img.avito.st/140x105/3155173217.jpg;Не удалось определить'],
    [
        '851166120;Hp ProBook 450 G2, core i5, 8gb, 6gb видео, обмен в Москве;28000;м.\xa0Площадь Ильича;Сегодня16:35;https://www.avito.ru/moskva/noutbuki/hp_probook_450_g2_core_i5_8gb_6gb_video_obmen_851166120;//76.img.avito.st/140x105/3066115876.jpg;Компания'],
    [
        '865671975;Lenovo Yoga 3 Pro Core M 4Gb 128SSD (гарантия) в Москве;44990;м.\xa0Динамо;Сегодня16:35;https://www.avito.ru/moskva/noutbuki/lenovo_yoga_3_pro_core_m_4gb_128ssd_garantiya_865671975;//81.img.avito.st/140x105/3140032381.jpg;None'],
    [
        '511313671;Новый Hp 250 15.6" 4-x яд 2.4ghzx4/6gb Озу/750gb в Москве;17000;м.\xa0Планерная;Сегодня16:35;https://www.avito.ru/moskva/noutbuki/novyy_hp_250_15.6_4-x_yad_2.4ghzx46gb_ozu750gb_511313671;//33.img.avito.st/140x105/1422217233.jpg;Компания'],
    [
        '869812087;Новый игровой ноутбук Acer Predator 17 G9-792-790G в Москве;150000;м.\xa0Багратионовская;Сегодня16:34;https://www.avito.ru/moskva/noutbuki/novyy_igrovoy_noutbuk_acer_predator_17_g9-792-790g_869812087;//41.img.avito.st/140x105/3155172041.jpg;Не удалось определить'],
    [
        '493460128;Lenovo b590, нов, 4-х яд 2.4ghzx4/4gb/500gb/gf720 в Москве;17500;м.\xa0Планерная;Сегодня16:34;https://www.avito.ru/moskva/noutbuki/lenovo_b590_nov_4-h_yad_2.4ghzx44gb500gbgf720_493460128;//14.img.avito.st/140x105/1353768314.jpg;Компания'],
    [
        '869811957;Нетбук Acer Aspire one в Москве;4500;м.\xa0Новогиреево;Сегодня16:34;https://www.avito.ru/moskva/noutbuki/netbuk_acer_aspire_one_869811957;//28.img.avito.st/140x105/3155176528.jpg;Не удалось определить'],
    [
        '869811935;MacBook Core2duo Fhd экран большой в Москве;18000;м.\xa0Петровско-Разумовская;Сегодня16:34;https://www.avito.ru/moskva/noutbuki/macbook_core2duo_fhd_ekran_bolshoy_869811935;//06.img.avito.st/140x105/3155175306.jpg;Не удалось определить'],
    [
        '869811898;Новый acer e15 17.3" 4-x яд Amd A4 6210/6gb/1Tb в Москве;19500;м.\xa0Планерная;Сегодня16:34;https://www.avito.ru/moskva/noutbuki/novyy_acer_e15_17.3_4-x_yad_amd_a4_62106gb1tb_869811898;//38.img.avito.st/140x105/3155171438.jpg;Компания'],
    [
        '770913555;Ноутбук делл pp44L 2-ядра в Москве;11600;м.\xa0Бибирево;Сегодня16:31;https://www.avito.ru/moskva/noutbuki/noutbuk_dell_pp44l_2-yadra_770913555;//00.img.avito.st/140x105/2598116300.jpg;Компания'],
    [
        '869810109;Продаю-обменяю samsung R525 500HDD в Москве;11000;м.\xa0Каширская;Сегодня16:31;https://www.avito.ru/moskva/noutbuki/prodayu-obmenyayu_samsung_r525_500hdd_869810109;//97.img.avito.st/140x105/3155169397.jpg;Не удалось определить'],
    [
        '869809780;Acer UltraBook 14" /256gb SSD/Intel Core i3 в Москве;21000;Не удалось определить;Сегодня16:31;https://www.avito.ru/moskva/noutbuki/acer_ultrabook_14_256gb_ssdintel_core_i3_869809780;//01.img.avito.st/140x105/3155167701.jpg;Не удалось определить'],
    [
        '869809706;Самсунг np-r150 в Москве;4900;м.\xa0Севастопольская;Сегодня16:30;https://www.avito.ru/moskva/noutbuki/samsung_np-r150_869809706;//39.img.avito.st/140x105/3155167239.jpg;Не удалось определить'],
    [
        '869809472;Ноутбук Toshiba Satellite R630-150 в Москве;10000;м.\xa0Измайловская;Сегодня16:30;https://www.avito.ru/moskva/noutbuki/noutbuk_toshiba_satellite_r630-150_869809472;//45.img.avito.st/140x105/3155159745.jpg;Не удалось определить'],
    [
        '851955472;Паккард - Игровой (шустрый) с 3 гигами в Москве;13300;м.\xa0Павелецкая;Сегодня16:29;https://www.avito.ru/moskva/noutbuki/pakkard_-_igrovoy_shustryy_s_3_gigami_851955472;//75.img.avito.st/140x105/3069628875.jpg;Магазин'],
    [
        '869807752;Apple MacBook Air 13 Early 2016 в Москве;16000;Не удалось определить;Сегодня16:27;https://www.avito.ru/moskva/noutbuki/apple_macbook_air_13_early_2016_869807752;//05.img.avito.st/140x105/3155160605.jpg;Не удалось определить'],
    [
        '869807658;Новый asus ROG G752VY-DH78K в Москве;183000;м.\xa0Фили;Сегодня16:27;https://www.avito.ru/moskva/noutbuki/novyy_asus_rog_g752vy-dh78k_869807658;//21.img.avito.st/140x105/3155149221.jpg;Компания'],
    [
        '869807649;Игровой ноутбук Acer Aspire 5742G в Москве;12000;м.\xa0Новогиреево;Сегодня16:27;https://www.avito.ru/moskva/noutbuki/igrovoy_noutbuk_acer_aspire_5742g_869807649;//12.img.avito.st/140x105/3155160612.jpg;Не удалось определить'],
    [
        '869807281;Macbook air 13 mid 13 в Москве;39000;м.\xa0Улица академика Янгеля;Сегодня16:26;https://www.avito.ru/moskva/noutbuki/macbook_air_13_mid_13_869807281;//76.img.avito.st/140x105/3155157976.jpg;Не удалось определить'],
    [
        '869807092;Asus s300ca core i7, 120 ssd, touch display в Москве;15000;м.\xa0Тимирязевская;Сегодня16:26;https://www.avito.ru/moskva/noutbuki/asus_s300ca_core_i7_120_ssd_touch_display_869807092;//11.img.avito.st/140x105/3155153111.jpg;Не удалось определить'],
    [
        '869806454;Ноутбуки разных брэндов в Москве;1000;м.\xa0Аэропорт;Сегодня16:25;https://www.avito.ru/moskva/noutbuki/noutbuki_raznyh_brendov_869806454;//34.img.avito.st/140x105/3155255634.jpg;Компания'],
    [
        '869805954;Классный ноутбук Asus X50N в Москве;5000;Компания;Сегодня16:24;https://www.avito.ru/moskva/noutbuki/klassnyy_noutbuk_asus_x50n_869805954;//14.img.avito.st/140x105/3155153514.jpg;Не удалось определить'],
    [
        '869805614;Apple MacBook Pro "Core 2 Duo" 2.4 15" 2.4 GHz 4g в Москве;15000;м.\xa0Савеловская;Сегодня16:24;https://www.avito.ru/moskva/noutbuki/apple_macbook_pro_core_2_duo_2.4_15_2.4_ghz_4g_869805614;//59.img.avito.st/140x105/3155150759.jpg;Не удалось определить'],
    [
        '851996377;Всем геймерам. Мощный c Core i7 процессорам в Москве;26999;м.\xa0Каширская;Сегодня16:23;https://www.avito.ru/moskva/noutbuki/vsem_geymeram._moschnyy_c_core_i7_protsessoram_851996377;//26.img.avito.st/140x105/3069859826.jpg;Компания'],
    [
        '869805410;Asus c DVD в Москве;3799;Не удалось определить;Сегодня16:23;https://www.avito.ru/moskva/noutbuki/asus_c_dvd_869805410;//82.img.avito.st/140x105/3155146082.jpg;Не удалось определить'],
    [
        '869805343;Ноутбук Asus x453m в Москве;15000;м.\xa0Семеновская;Сегодня16:23;https://www.avito.ru/moskva/noutbuki/noutbuk_asus_x453m_869805343;//73.img.avito.st/140x105/3155150673.jpg;Компания'],
    [
        '811197292;Toshiba в Москве;2500;м.\xa0Выхино;Сегодня16:23;https://www.avito.ru/moskva/noutbuki/toshiba_811197292;//50.img.avito.st/140x105/2833184550.jpg;Не удалось определить'],
    [
        '367436715;Игровой ноутбук 15.6" Acer 5000, 8-мь core i7 с 8g в Москве;24990;м.\xa0Планерная;Сегодня16:23;https://www.avito.ru/moskva/noutbuki/igrovoy_noutbuk_15.6_acer_5000_8-m_core_i7_s_8g_367436715;//94.img.avito.st/140x105/924843294.jpg;Компания'],
    [
        '687161200;Lenovo G580, Игровой ноутбук 2.4ghzx4/4gb Озу/500g в Москве;16000;м.\xa0Планерная;Сегодня16:22;https://www.avito.ru/moskva/noutbuki/lenovo_g580_igrovoy_noutbuk_2.4ghzx44gb_ozu500g_687161200;Нет картинки;Компания'],
    [
        '869803797;8 ядерный, очень мощный Acer в Москве;22000;Не удалось определить;Сегодня16:21;https://www.avito.ru/moskva/noutbuki/8_yadernyy_ochen_moschnyy_acer_869803797;//20.img.avito.st/140x105/3155145420.jpg;Не удалось определить'],
    [
        '837863153;Продам ноутбук MSI i5/16 Gb/2 Gb/SSD 120GB+ HDD 1T в Москве;57000;м.\xa0Юго-Западная;Сегодня16:21;https://www.avito.ru/moskva/noutbuki/prodam_noutbuk_msi_i516_gb2_gbssd_120gb_hdd_1t_837863153;//37.img.avito.st/140x105/2987016337.jpg;Не удалось определить'],
    [
        '834866501;В коробке Белый Sony Vaio core i3/4gb/3gb GeForce в Москве;17500;м.\xa0Планерная;Сегодня16:21;https://www.avito.ru/moskva/noutbuki/v_korobke_belyy_sony_vaio_core_i34gb3gb_geforce_834866501;//26.img.avito.st/140x105/2969443626.jpg;Компания'],
    [
        '869803502;Sony PCG-6GHP в Москве;5500;м.\xa0Бибирево;Сегодня16:20;https://www.avito.ru/moskva/noutbuki/sony_pcg-6ghp_869803502;//68.img.avito.st/140x105/3155143368.jpg;Не удалось определить'],
    [
        '869803351;Продам ноутбук BenQ Joybook A52-R02 в Москве;5000;м.\xa0Савеловская;Сегодня16:20;https://www.avito.ru/moskva/noutbuki/prodam_noutbuk_benq_joybook_a52-r02_869803351;//64.img.avito.st/140x105/3155132664.jpg;Не удалось определить'],
    [
        '734738080;Продаются запчасти для ноутбука Acer Aspire 5552G в Москве;2000;м.\xa0Дмитровская;Сегодня16:20;https://www.avito.ru/moskva/noutbuki/prodayutsya_zapchasti_dlya_noutbuka_acer_aspire_5552g_734738080;//70.img.avito.st/140x105/2310647870.jpg;Компания'],
    [
        '869802827;Lenovo S10-3S в Москве;7200;Не удалось определить;Сегодня16:19;https://www.avito.ru/moskva/noutbuki/lenovo_s10-3s_869802827;//48.img.avito.st/140x105/3155142648.jpg;Не удалось определить'],
    [
        '806360183;Acer Aspire 5738PZG-443G25MI под запчасти в Москве;3200;м.\xa0Калужская;Сегодня16:18;https://www.avito.ru/moskva/noutbuki/acer_aspire_5738pzg-443g25mi_pod_zapchasti_806360183;//34.img.avito.st/140x105/2804195434.jpg;Не удалось определить']]
proxy = PROXY

'''старые вызовы'''


def get_html(url, proxy):
    retry_conn = 5
    # запрашиваем страницу
    try:
        response = requests.get(url, proxies=proxy)
    except Exception:
        print('\n Ошибка соединения. Пробую снова...')
        while not retry_conn:
            retry_conn = - 1
            print('\n Ошибка соединения. Пробую снова...')
            raise
        print('Соединение не удалось. \n Сохраняю проект.')
        save(project, 'storage.csv')
        print('Проект сохранён.')
        input('Для выхода нажмите Enter')
        exit()

    # тут я не понял что делаем, но это нужно для дальнейших действий
    return html.document_fromstring(response.content)


def parsing_page(doc):
    items_table = []
    items = doc.cssselect('div.js-catalog_after-ads .item_table')  # отсекаем всЁ до нужного нам раздела

    for item in items:
        id_item = (item.get('id')[1:])  # узнаём ID товара
        href = 'https://www.avito.ru' + item.cssselect('div.description h3 a')[0].get('href')  # узнаём ссылку на товар
        title = item.cssselect('div.description h3 a')[0].get('title')  # узнаём заголовок объявления

        try:
            src = item.cssselect('div.b-photo a img')[0].get('src')  # узнаём ссылку на картинку для объявления
        except:
            src = 'Нет картинки'

        try:
            price = (item.cssselect('div.description .about')[0].text)  # узнаём цену товара
            price = price.replace('\n', '').replace('руб.', '').replace(' ', '')  # отсекаем лишние символы
        except:
            price = str('0')

        try:
            podrazdel = str(item.cssselect('div.description > div.data > p:nth-child(1)')[0].text)
        except:
            podrazdel = 'Не удалось определить'  # узнаём подраздел или город

        try:
            city = item.cssselect('div.description div.data p:nth-child(2)')[0].text
            # узнаём город в котором продаётся товар
        except:
            city = podrazdel  # город не обнаружен, значит раздела нет,
            # а в переменную для раздела записан город, поэтому присваеваем
            # городу  значение, которое получил раздел,а
            podrazdel = 'Не удалось определить'  # разделу приваеваем значение 'Не удалось определить'

        try:
            # date_item = ''
            date_item = str((item.cssselect('.item_table .data .date')[0].text).replace('\n', '').replace(' ', ''))
            # узнаём дату публикации
            if date_item == None: date_item = 'Не удалось определить'
        except:
            date_item = 'Не удалось определить'

        items_table.append([id_item
                            + ';' + title
                            + ';' + price
                            + ';' + city
                            + ';' + date_item
                            + ';' + href
                            + ';' + src
                            + ';' + podrazdel
                            ])
    return items_table


'''новые вызовы'''


def get_config(conf_file):
    config = ConfigParser()
    config.read(conf_file)
    # http_proxy = config.get('general', 'http_proxy')
    # https_proxy = config.get('general', 'https_proxy')
    url = config.get('general', 'url')
    pages = config.get('general', 'pages')
    return url, int(pages)


def parsing_avito_page(url, proxy):
    items_table = []

    retry_conn = 5
    # запрашиваем страницу
    try:
        response = requests.get(url, proxies=proxy)
    except Exception:
        print('\n Ошибка соединения. Пробую снова...')
        while not retry_conn:
            retry_conn = - 1
            print('\n Ошибка соединения. Пробую снова...')
            raise
        print('Соединение не удалось. \n Сохраняю проект.')
        save(project, 'storage.csv')
        print('Проект сохранён.')
        input('Для выхода нажмите Enter')
        exit()

    # тут я не понял что делаем, но это нужно для дальнейших действий
    doc = html.document_fromstring(response.content)
    items = doc.cssselect('div.js-catalog_after-ads .item_table')  # отсекаем всЁ до нужного нам раздела

    for item in items:
        id_item = (item.get('id')[1:])  # узнаём ID товара
        href = 'https://www.avito.ru' + item.cssselect('div.description h3 a')[0].get('href')  # узнаём ссылку на товар
        title = item.cssselect('div.description h3 a')[0].get('title')  # узнаём заголовок объявления

        try:
            src = item.cssselect('div.b-photo a img')[0].get('src')  # узнаём ссылку на картинку для объявления
        except:
            src = 'Нет картинки'

        try:
            price = (item.cssselect('div.description .about')[0].text)  # узнаём цену товара
            price = price.replace('\n', '').replace('руб.', '').replace(' ', '')  # отсекаем лишние символы
        except:
            price = str('0')

        try:
            podrazdel = str(item.cssselect('div.description > div.data > p:nth-child(1)')[0].text)
        except:
            podrazdel = 'Не удалось определить'  # узнаём подраздел или город

        try:
            city = item.cssselect('div.description div.data p:nth-child(2)')[0].text
            # узнаём город в котором продаётся товар
        except:
            city = podrazdel  # город не обнаружен, значит раздела нет,
            # а в переменную для раздела записан город, поэтому присваеваем
            # городу  значение, которое получил раздел,а
            podrazdel = 'Не удалось определить'  # разделу приваеваем значение 'Не удалось определить'

        try:
            # date_item = ''
            date_item = str((item.cssselect('.item_table .data .date')[0].text).replace('\n', '').replace(' ', ''))
            # узнаём дату публикации
            if date_item == None: date_item = 'Не удалось определить'
        except:
            date_item = 'Не удалось определить'

        items_table.append([id_item
                            + ';' + title
                            + ';' + price
                            + ';' + city
                            + ';' + date_item
                            + ';' + href
                            + ';' + src
                            + ';' + podrazdel
                            ])
    return items_table


def get_next_url(url, count):
    try:
        index_sign = url.index('?')
    except ValueError:
        return url + '?p=' + str(count)
    return (url[:index_sign] + '?p=' + str(count) + '&' + url[(index_sign + 1):])


def save(projects, path):
    # Переименовываем старый файл
    if os.path.exists(path): os.replace(path, path + '.bak')
    try:
        with open(path, 'w') as csvfile:
            writer = csv.writer(csvfile, quoting=csv.QUOTE_NONE, delimiter='|', quotechar='')

            writer.writerow(['ID: ; Заголовок:; Цена:; Город размещения:; Дата размещения; Ссылка на товар: ;'
                             ' Ссылка на изображение:; Подраздел:'])
            for item in projects:
                writer.writerow(item)
        print('Файл успешно сохранён!')
    except IOError:
        print('Не могу открыть файл storage.csv. Файл может быть заблокирован другой программой')
        if input('Закройте все программы которые могут использовать файл и нажмите Enter чтобы поробовать еще раз.'
                 '(N , Enter - для отмены)') == None: raise
        print('Что-то пошло не так. сохранение не удалось')
    except:
        print('Что-то пошло не так. сохранение не удалось')


def xls_write(project, full_filename):
    # принимает словарь где ключ это ID товара на авито и файл который необходимо дополнить информацией
    try:
        wb = load_workbook(full_filename)
        ws = wb.active
        rows = len(project)
        for row in range(rows):
            cols = len(project[0])
            for col in cols:
                ws.col(row=row, column=col).value = project[row][col]
        wb.save()
        print('Файл успешно сохранён!')
    except IOError:
        print('Не могу открыть файл storage.csv. Файл может быть заблокирован другой программой')
        if input('Закройте все программы которые могут использовать файл и нажмите Enter чтобы поробовать еще раз.'
                 '(N , Enter - для отмены)') == None: raise
        print('Что-то пошло не так. сохранение не удалось')
    except:
        print('Что-то пошло не так. Сохранение не удалось')


def read_xls(full_filename):
    table = []
    wb = load_workbook(full_filename)
    # print(wb)
    ws = wb.active
    for row in ws.iter_rows():
        table_row = []
        for col in row:
            table_row.append(col.value)
        table.append(table_row)
    # wb.close()
    return table


def list_to_dict(project_list):
    project_dict = {}
    for row in project_list:
        project_dict[row[0]] = row[1:]
    return project_dict


def get_table(url, proxy, pages):
    for i in range(1, pages + 1):
        page = []
        if i == 1:
            project.extend(parsing_avito_page(url, proxy))
            print(i, '-ая страница, по ссылке ', url)
            continue
        next_url = get_next_url(url, i)
        page = parsing_avito_page(next_url, proxy)
        print(i, '-ая страница, по ссылке ', next_url)
        if not page:
            print("Страница пустая. Заканчиваем", '\n')
            break
        project.extend(page)
    return project


def dict_to_list(dict):
    dictlist = []
    for key, value in dict:
        dictlist.append([dict.items(key, value)])
    return dictlist


def main():
    # url, pages = read_conf_file(conf_file)
    # print("Page 1 - getting on URL", url, end='')
    # html_page = get_html(url)
    # print(" ... parsing")
    # page = parsing_avito_page(html_page)
    # project.extend(page)
    # print("Done!")
    # for i in range(2, pages + 1):
    #     page.clear()
    #     next_url = get_next_url(url, i)
    #     print("Page", i, "getting on URL", next_url, end='')
    #     html_page = get_html(next_url)
    #     print(" ... parsing")
    #     page = parsing_avito_page(html_page)
    #     if not page:
    #         print("Данные больше не найдены, страница пустая, значит страницы по поиску кончились. Заканчиваем", '\n')
    #         break
    #     project.extend(page)
    #     print("Done!")
    # save(project, 'storage.csv')
    # inp = input("Нажмите Enter:")
    path = FPATH
    url, pages = get_config(conf_file)  # получение полного конфига
    print("Прочитали и спарсили конфиг", conf_file)
    old_project = read_xls(path)  # чтение спарсенных ранее страниц из файла
    print('Прочитали файл', path)
    print("Вот он \n", old_project)
    old_table = list_to_dict(old_project)  # преобразование полученной таблицы в словарь где ID товара является ключом
    print("Преобразовали список в словарь.")
    print("Вот он \n", old_table)
    print("Получаем страницы по указанной ссылке", url)
    new_project = get_table(url, proxy, pages)  # получение таблицы которую необходимо добавить в наш файл
    print("Получили все нужные страницы и вытащили из них данные")
    print("Вот они \n", new_project)
    # преобразование таблицы в словарь
    new_table = list_to_dict(new_project)
    print("Преобразовали полученный список в словарь")
    print("Вот он \n", new_table)
    # совмещение обоих словарей
    old_table.update(new_table)
    print("Совместили словари, исключив повторения")
    print("Вот он \n", old_table)
    # преобразование словаря в таблицу
    new_project = dict_to_list(old_table)
    print("Преобразовали словарь обратно в список")
    # запись таблицы обратно в файл
    xls_write(new_project, path)
    print("Записали изменения в файл", path)
    input("Для выхода нажмите Enter")


if __name__ == '__main__':
    main()
